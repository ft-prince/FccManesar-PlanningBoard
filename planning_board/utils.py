# utils.py - Enhanced Excel Processing
import openpyxl
from datetime import datetime, time, timedelta
import re
from django.utils import timezone
from .models import (
    PlanningBoard, ProductionLine, TomorrowPlan, NextDayPlan,
    CriticalPartStatus, AFMPlan, SPDPlan, OtherInformation
)

class ExcelProcessor:
    """Enhanced Excel file processor for planning board data"""
    
    def __init__(self, file_path, planning_board):
        self.file_path = file_path
        self.planning_board = planning_board
        self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        self.worksheet = self.workbook.active
    
    def process_excel(self):
        """Main method to process the entire Excel file"""
        try:
            self.extract_basic_info()
            self.extract_production_lines()
            self.extract_tomorrow_plans()
            self.extract_next_day_plans()
            self.extract_critical_parts()
            self.extract_afm_plans()
            self.extract_spd_plans()
            self.extract_other_information()
            return True, "Excel file processed successfully"
        except Exception as e:
            return False, f"Error processing Excel: {str(e)}"
    
    def extract_basic_info(self):
        """Extract basic planning board information"""
        # Extract meeting time from cell B2
        meeting_cell = self.worksheet['B2'].value
        if meeting_cell and isinstance(meeting_cell, str):
            time_match = re.search(r'(\d{1,2}):(\d{2})', meeting_cell)
            if time_match:
                hour, minute = int(time_match.group(1)), int(time_match.group(2))
                self.planning_board.meeting_time = time(hour, minute)
        
        # Extract dates from row 3
        date_cells = ['C3', 'T3', 'Y3']  # Today, Tomorrow, Next day
        dates = []
        
        for cell_addr in date_cells:
            cell_value = self.worksheet[cell_addr].value
            if cell_value:
                # Try to extract date from string
                date_obj = self.parse_date_from_cell(cell_value)
                if date_obj:
                    dates.append(date_obj)
        
        # Update planning board dates if found
        if len(dates) >= 1:
            self.planning_board.today_date = dates[0]
        if len(dates) >= 2:
            self.planning_board.tomorrow_date = dates[1]
        if len(dates) >= 3:
            self.planning_board.next_day_date = dates[2]
        
        self.planning_board.save()
    
    def parse_date_from_cell(self, cell_value):
        """Parse date from various cell formats"""
        if isinstance(cell_value, datetime):
            return cell_value.date()
        
        if isinstance(cell_value, str):
            # Remove "DATE:-" prefix if present
            date_str = cell_value.replace('DATE:-', '').strip()
            
            # Try different date formats
            date_formats = ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%m/%d/%Y']
            for fmt in date_formats:
                try:
                    return datetime.strptime(date_str, fmt).date()
                except ValueError:
                    continue
        
        return None
    
    def extract_production_lines(self):
        """Extract production line data from Excel"""
        # Production line names and their approximate row positions
        line_configs = [
            {'name': 'CLUTCH ASSY LINE-1', 'row': 7},
            {'name': 'CLUTCH ASSY LINE-2', 'row': 13},
            {'name': 'PULLEY ASSY LINE-1', 'row': 19},
            {'name': 'FMD/FFD', 'row': 22},
            {'name': 'NEW BUSSINESS', 'row': 26},
        ]
        
        for config in line_configs:
            self.process_production_line(config['name'], config['row'])
    
    def process_production_line(self, line_name, start_row):
        """Process individual production line data"""
        # Check if line exists in Excel
        line_cell = self.worksheet.cell(row=start_row, column=2).value
        if not line_cell or line_name not in str(line_cell):
            return
        
        # Extract shift data
        production_line = ProductionLine.objects.create(
            planning_board=self.planning_board,
            line_number=line_name
        )
        
        # Data rows can span multiple rows, so we'll look in the range
        for row_offset in range(6):  # Check up to 6 rows
            current_row = start_row + row_offset
            
            # A Shift data (columns C-H)
            if not production_line.a_shift_model:
                production_line.a_shift_model = self.get_cell_value(current_row, 3)
                production_line.a_shift_plan = self.get_numeric_value(current_row, 4)
                production_line.a_shift_actual = self.get_numeric_value(current_row, 5)
                production_line.a_shift_plan_change = self.get_numeric_value(current_row, 6)
                production_line.a_shift_time = self.get_time_value(current_row, 7)
                production_line.a_shift_remarks = self.get_cell_value(current_row, 8)
            
            # B Shift data (columns I-N)
            if not production_line.b_shift_model:
                production_line.b_shift_model = self.get_cell_value(current_row, 9)
                production_line.b_shift_plan = self.get_numeric_value(current_row, 10)
                production_line.b_shift_actual = self.get_numeric_value(current_row, 11)
                production_line.b_shift_plan_change = self.get_numeric_value(current_row, 12)
                production_line.b_shift_time = self.get_time_value(current_row, 13)
                production_line.b_shift_remarks = self.get_cell_value(current_row, 14)
            
            # C Shift data (columns O-S)
            if not production_line.c_shift_model:
                production_line.c_shift_model = self.get_cell_value(current_row, 15)
                production_line.c_shift_plan = self.get_numeric_value(current_row, 16)
                production_line.c_shift_actual = self.get_numeric_value(current_row, 17)
                production_line.c_shift_plan_change = self.get_numeric_value(current_row, 18)
                production_line.c_shift_remarks = self.get_cell_value(current_row, 19)
        
        production_line.save()
    
    def extract_tomorrow_plans(self):
        """Extract tomorrow assembly plans"""
        # Tomorrow plans are around columns T-V, starting from row 6
        start_col = 20  # Column T
        start_row = 6
        
        for row in range(start_row, start_row + 10):  # Check multiple rows
            model = self.get_cell_value(row, start_col)
            if model and model.strip():
                TomorrowPlan.objects.create(
                    planning_board=self.planning_board,
                    model=model,
                    a_shift=self.get_numeric_value(row, start_col + 1),
                    b_shift=self.get_numeric_value(row, start_col + 2),
                    c_shift=self.get_numeric_value(row, start_col + 3),
                    remarks=self.get_cell_value(row, start_col + 4)
                )
    
    def extract_next_day_plans(self):
        """Extract next day assembly plans"""
        # Next day plans are around columns Y-AA, starting from row 6
        start_col = 25  # Column Y
        start_row = 6
        
        for row in range(start_row, start_row + 10):
            model = self.get_cell_value(row, start_col)
            if model and model.strip():
                NextDayPlan.objects.create(
                    planning_board=self.planning_board,
                    model=model,
                    a_shift=self.get_numeric_value(row, start_col + 1),
                    b_shift=self.get_numeric_value(row, start_col + 2),
                    c_shift=self.get_numeric_value(row, start_col + 3),
                    remarks=self.get_cell_value(row, start_col + 4)
                )
    
    def extract_critical_parts(self):
        """Extract critical part status data"""
        # Critical parts start around row 31
        start_row = 31
        start_col = 2  # Column B
        
        for row in range(start_row, start_row + 10):
            part_name = self.get_cell_value(row, start_col)
            if part_name and part_name.strip():
                receiving_time_str = self.get_cell_value(row, start_col + 3)
                receiving_time = None
                if receiving_time_str:
                    receiving_time = self.parse_datetime_from_cell(receiving_time_str)
                
                CriticalPartStatus.objects.create(
                    planning_board=self.planning_board,
                    part_name=part_name,
                    supplier=self.get_cell_value(row, start_col + 1) or '',
                    plan_qty=self.get_numeric_value(row, start_col + 2) or 0,
                    receiving_time=receiving_time,
                    remarks=self.get_cell_value(row, start_col + 4) or ''
                )
    
    def extract_afm_plans(self):
        """Extract AFM plan data"""
        # AFM plans are in columns around F-I, starting from row 31
        start_row = 31
        
        # FCIN plans
        for row in range(start_row, start_row + 10):
            part_name = self.get_cell_value(row, 6)  # Column F
            if part_name and part_name.strip():
                AFMPlan.objects.create(
                    planning_board=self.planning_board,
                    plan_type='FCIN',
                    part_name=part_name,
                    part_number=self.get_cell_value(row, 7) or '',
                    plan_qty=self.get_numeric_value(row, 8) or 0,
                    remarks=self.get_cell_value(row, 9) or ''
                )
    
    def extract_spd_plans(self):
        """Extract SPD plan customer-wise data"""
        # SPD plans start around column J, row 31
        start_row = 31
        customers = ['MSIL', 'HMSI', 'IYM', 'HMCL']
        
        col_offset = 10  # Starting from column J
        for i, customer in enumerate(customers):
            current_col = col_offset + (i * 4)  # Each customer has 4 columns
            
            for row in range(start_row, start_row + 10):
                part_name = self.get_cell_value(row, current_col)
                if part_name and part_name.strip():
                    SPDPlan.objects.create(
                        planning_board=self.planning_board,
                        customer=customer,
                        part_name=part_name,
                        part_number=self.get_cell_value(row, current_col + 1) or '',
                        plan_qty=self.get_numeric_value(row, current_col + 2) or 0,
                        remarks=self.get_cell_value(row, current_col + 3) or ''
                    )
    
    def extract_other_information(self):
        """Extract other information data"""
        # Other information starts around column Z, row 31
        start_row = 31
        start_col = 26  # Column Z
        
        for row in range(start_row, start_row + 10):
            part_name = self.get_cell_value(row, start_col)
            if part_name and part_name.strip():
                target_date_str = self.get_cell_value(row, start_col + 2)
                target_date = None
                if target_date_str:
                    target_date = self.parse_date_from_cell(target_date_str)
                
                if target_date:
                    OtherInformation.objects.create(
                        planning_board=self.planning_board,
                        part_name=part_name,
                        qty=self.get_numeric_value(row, start_col + 1) or 0,
                        target_date=target_date,
                        remarks=''
                    )
    
    def get_cell_value(self, row, col):
        """Get cell value as string"""
        cell = self.worksheet.cell(row=row, column=col)
        return str(cell.value).strip() if cell.value else ''
    
    def get_numeric_value(self, row, col):
        """Get cell value as number"""
        cell = self.worksheet.cell(row=row, column=col)
        if cell.value is None:
            return None
        try:
            return int(float(cell.value))
        except (ValueError, TypeError):
            return None
    
    def get_time_value(self, row, col):
        """Get cell value as time"""
        cell = self.worksheet.cell(row=row, column=col)
        if isinstance(cell.value, time):
            return cell.value
        if isinstance(cell.value, datetime):
            return cell.value.time()
        if isinstance(cell.value, str):
            time_match = re.search(r'(\d{1,2}):(\d{2})', cell.value)
            if time_match:
                hour, minute = int(time_match.group(1)), int(time_match.group(2))
                return time(hour, minute)
        return None
    
    def parse_datetime_from_cell(self, cell_value):
        """Parse datetime from cell value"""
        if isinstance(cell_value, datetime):
            return timezone.make_aware(cell_value)
        
        if isinstance(cell_value, str):
            # Try different datetime formats
            datetime_formats = [
                '%d/%m/%Y %H:%M',
                '%Y-%m-%d %H:%M',
                '%d-%m-%Y %H:%M'
            ]
            for fmt in datetime_formats:
                try:
                    dt = datetime.strptime(cell_value.strip(), fmt)
                    return timezone.make_aware(dt)
                except ValueError:
                    continue
        
        return None

# Updated views.py process_excel_file function
def process_excel_file(file, board):
    """Process uploaded Excel file and populate database"""
    processor = ExcelProcessor(file, board)
    success, message = processor.process_excel()
    if not success:
        raise Exception(message)
    return True