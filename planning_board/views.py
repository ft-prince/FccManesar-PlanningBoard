# views.py - Complete fixed version with proper Excel processing
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.utils import timezone
from django.db.models import Q

from datetime import datetime, timedelta
import openpyxl
import io
from .models import (
    PlanningBoard, ProductionLine, TomorrowPlan, NextDayPlan,
    CriticalPartStatus, AFMPlan, SPDPlan, OtherInformation, ExcelUpload
)
from .forms import (
    PlanningBoardForm, ExcelUploadForm, ProductionLineFormSet,
    TomorrowPlanFormSet, NextDayPlanFormSet, CriticalPartStatusFormSet,
    AFMPlanFormSet, SPDPlanFormSet, OtherInformationFormSet
)

@login_required
def planning_board_list(request):
    """List all planning boards"""
    boards = PlanningBoard.objects.filter(created_by=request.user).order_by('-created_at')
    return render(request, 'planning_board/list.html', {'boards': boards})

@login_required
def planning_board_detail(request, pk):
    """View a specific planning board"""
    board = get_object_or_404(PlanningBoard, pk=pk, created_by=request.user)
    return render(request, 'planning_board/detail.html', {'board': board})

@login_required
def planning_board_create(request):
    """Create a new planning board"""
    if request.method == 'POST':
        form = PlanningBoardForm(request.POST)
        if form.is_valid():
            board = form.save(commit=False)
            board.created_by = request.user
            board.save()
            messages.success(request, 'Planning board created successfully!')
            return redirect('planning_board:edit', pk=board.pk)
    else:
        # Set default dates
        today = timezone.now().date()
        form = PlanningBoardForm(initial={
            'today_date': today,
            'tomorrow_date': today + timedelta(days=1),
            'next_day_date': today + timedelta(days=2),
        })
    
    return render(request, 'planning_board/create.html', {'form': form})

@login_required
def planning_board_edit(request, pk):
    """Edit a planning board with all related data"""
    board = get_object_or_404(PlanningBoard, pk=pk, created_by=request.user)
    
    if request.method == 'POST':
        form = PlanningBoardForm(request.POST, instance=board)
        production_formset = ProductionLineFormSet(request.POST, instance=board)
        tomorrow_formset = TomorrowPlanFormSet(request.POST, instance=board)
        next_day_formset = NextDayPlanFormSet(request.POST, instance=board)
        critical_formset = CriticalPartStatusFormSet(request.POST, instance=board)
        afm_formset = AFMPlanFormSet(request.POST, instance=board)
        spd_formset = SPDPlanFormSet(request.POST, instance=board)
        other_formset = OtherInformationFormSet(request.POST, instance=board)
        
        if (form.is_valid() and production_formset.is_valid() and 
            tomorrow_formset.is_valid() and next_day_formset.is_valid() and
            critical_formset.is_valid() and afm_formset.is_valid() and
            spd_formset.is_valid() and other_formset.is_valid()):
            
            form.save()
            production_formset.save()
            tomorrow_formset.save()
            next_day_formset.save()
            critical_formset.save()
            afm_formset.save()
            spd_formset.save()
            other_formset.save()
            
            messages.success(request, 'Planning board updated successfully!')
            return redirect('planning_board:detail', pk=board.pk)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PlanningBoardForm(instance=board)
        production_formset = ProductionLineFormSet(instance=board)
        tomorrow_formset = TomorrowPlanFormSet(instance=board)
        next_day_formset = NextDayPlanFormSet(instance=board)
        critical_formset = CriticalPartStatusFormSet(instance=board)
        afm_formset = AFMPlanFormSet(instance=board)
        spd_formset = SPDPlanFormSet(instance=board)
        other_formset = OtherInformationFormSet(instance=board)
    
    context = {
        'form': form,
        'board': board,
        'production_formset': production_formset,
        'tomorrow_formset': tomorrow_formset,
        'next_day_formset': next_day_formset,
        'critical_formset': critical_formset,
        'afm_formset': afm_formset,
        'spd_formset': spd_formset,
        'other_formset': other_formset,
    }
    
    return render(request, 'planning_board/edit.html', context)

@login_required
def planning_board_delete(request, pk):
    """Delete a planning board"""
    board = get_object_or_404(PlanningBoard, pk=pk, created_by=request.user)
    
    if request.method == 'POST':
        board.delete()
        messages.success(request, 'Planning board deleted successfully!')
        return redirect('planning_board:list')
    
    return render(request, 'planning_board/delete.html', {'board': board})

@login_required
def excel_upload(request):
    """Upload and process Excel file"""
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            # Create planning board first
            today = timezone.now().date()
            board = PlanningBoard.objects.create(
                created_by=request.user,
                today_date=today,
                tomorrow_date=today + timedelta(days=1),
                next_day_date=today + timedelta(days=2),
            )
            
            # Save the upload record
            upload = form.save(commit=False)
            upload.planning_board = board
            upload.uploaded_by = request.user
            upload.save()
            
            # Process the Excel file
            try:
                success = process_excel_file(upload.file.path, board)
                if success:
                    upload.processed = True
                    upload.save()
                    messages.success(request, f'Excel file uploaded and processed successfully! Planning board created with ID: {board.id}')
                    return redirect('planning_board:detail', pk=board.pk)
                else:
                    board.delete()
                    messages.error(request, 'Error processing Excel file. Please check the file format.')
            except Exception as e:
                board.delete()  # Clean up if processing fails
                messages.error(request, f'Error processing Excel file: {str(e)}')
    else:
        form = ExcelUploadForm()
    
    return render(request, 'planning_board/excel_upload.html', {'form': form})

def process_excel_file(file_path, board):
    """Process uploaded Excel file and populate database - enhanced version"""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        worksheet = workbook.active
        
        print(f"Processing Excel file with {worksheet.max_row} rows and {worksheet.max_column} columns")
        
        # Debug the Excel structure first
        debug_excel_structure(worksheet)
        
        # Extract basic information
        extract_basic_info(worksheet, board)
        
        # Extract production lines data
        extract_production_lines(worksheet, board)
        
        # Extract future planning data with fixed approach
        extract_future_plans_fixed(worksheet, board)
        
        # Extract additional sections using enhanced approach
        extract_additional_sections(worksheet, board)
        
        return True
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False

def debug_excel_structure(worksheet):
    """Debug function to understand Excel layout"""
    try:
        print("=== DEBUGGING EXCEL STRUCTURE ===")
        
        # Look for plan headers specifically
        for row in range(1, 15):
            row_content = []
            for col in range(1, 35):
                cell_value = get_cell_value(worksheet, row, col)
                if cell_value and any(keyword in cell_value.upper() for keyword in ['TOMORROW', 'NEXT', 'DAY', 'PLAN', 'MODEL', 'SHIFT']):
                    col_letter = chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"
                    row_content.append(f"Col {col}({col_letter}): '{cell_value}'")
            
            if row_content:
                print(f"Row {row}: {' | '.join(row_content)}")
        
        print("=== END DEBUG ===")
    except Exception as e:
        print(f"Error in debug: {e}")

def extract_basic_info(worksheet, board):
    """Extract basic information from Excel"""
    try:
        # Extract meeting time from B2
        meeting_cell = worksheet['B2'].value
        if meeting_cell and 'TIME' in str(meeting_cell).upper():
            # Try to extract time
            import re
            time_match = re.search(r'(\d{1,2}):(\d{2})', str(meeting_cell))
            if time_match:
                from datetime import time
                hour, minute = int(time_match.group(1)), int(time_match.group(2))
                board.meeting_time = time(hour, minute)
        
        # Update title from C2
        title_cell = worksheet['C2'].value
        if title_cell:
            board.title = str(title_cell)
        
        board.save()
    except Exception as e:
        print(f"Error extracting basic info: {e}")

def extract_production_lines(worksheet, board):
    """Extract production line data from specific rows - improved version"""
    try:
        # Define production lines and their starting row positions
        line_configs = [
            {'name': 'PULLEY ASSY LINE-1', 'start_row': 7, 'max_rows': 3},    # 3 items
            {'name': 'CLUTCH ASSY LINE-3', 'start_row': 10, 'max_rows': 16},  # 17 items
            {'name': 'CLUTCH ASSY LINE-2', 'start_row': 26, 'max_rows': 17},  # 16 items
            {'name': 'FMD/FFD', 'start_row': 43, 'max_rows': 4},              # 3 items
        ]

        
        for config in line_configs:
            extract_single_production_line(worksheet, board, config)
                
    except Exception as e:
        print(f"Error extracting production lines: {e}")

def extract_single_production_line(worksheet, board, config):
    """Extract data for a single production line with multiple entries"""
    try:
        line_name = config['name']
        start_row = config['start_row']
        max_rows = config.get('max_rows', 5)
        
        # Track all data entries for this production line
        line_entries = []
        
        # Scan through all possible rows for this production line
        for row_offset in range(max_rows):
            current_row = start_row + row_offset
            
            # Check if this row has any meaningful data
            row_data = extract_row_data(worksheet, current_row)
            
            if has_meaningful_data(row_data):
                line_entries.append(row_data)
        
        # Create production line entries
        if line_entries:
            create_production_line_entries(board, line_name, line_entries)
        else:
            # Create empty production line if no data found
            ProductionLine.objects.create(
                planning_board=board,
                line_number=line_name
            )
            
    except Exception as e:
        print(f"Error extracting line {config['name']}: {e}")

def extract_row_data(worksheet, row):
    """Extract all data from a single row"""
    try:
        return {
            # A Shift data
            'a_shift': {
                'model': get_cell_value(worksheet, row, 3),       # Column C
                'plan': get_numeric_value(worksheet, row, 4),     # Column D
                'plan_change': get_numeric_value(worksheet, row, 5), # Column F
                'actual': get_numeric_value(worksheet, row, 6),   # Column E
                'time': get_time_value(worksheet, row, 7),        # Column G
                'remarks': get_cell_value(worksheet, row, 8),     # Column H
            },
            # B Shift data
            'b_shift': {
                'model': get_cell_value(worksheet, row, 9),       # Column I
                'plan': get_numeric_value(worksheet, row, 10),    # Column J
                'plan_change': get_numeric_value(worksheet, row, 11), # Column L
                'actual': get_numeric_value(worksheet, row, 12),  # Column K
                'time': get_time_value(worksheet, row, 13),       # Column M
                'remarks': get_cell_value(worksheet, row, 14),    # Column N
            },
            # C Shift data
            'c_shift': {
                'model': get_cell_value(worksheet, row, 15),      # Column O
                'plan': get_numeric_value(worksheet, row, 16),    # Column P
                'plan_change': get_numeric_value(worksheet, row, 17), # Column R
                'actual': get_numeric_value(worksheet, row, 18),  # Column Q
                'remarks': get_cell_value(worksheet, row, 20),    # Column S
            }
        }
    except Exception as e:
        print(f"Error extracting row {row}: {e}")
        return None
    
def has_meaningful_data(row_data):
    """Check if a row contains meaningful data"""
    if not row_data:
        return False
    
    # Check if any shift has a model name (indicating actual data)
    for shift in ['a_shift', 'b_shift', 'c_shift']:
        if row_data[shift]['model'] and row_data[shift]['model'].strip():
            # Exclude header rows
            model = row_data[shift]['model'].strip().upper()
            if model not in ['MODEL', 'SHIFT', 'LINE', 'NO.', '']:
                return True
    
    return False

def create_production_line_entries(board, line_name, line_entries):
    """Create production line database entries from extracted data"""
    try:
        entry_count = 0
        
        for entry_data in line_entries:
            # Check which shifts have data
            shifts_with_data = []
            for shift_name in ['a_shift', 'b_shift', 'c_shift']:
                if entry_data[shift_name]['model'] and entry_data[shift_name]['model'].strip():
                    shifts_with_data.append(shift_name)
            
            if shifts_with_data:
                entry_count += 1
                # Create line name with entry number if multiple entries
                display_name = f"{line_name}" if entry_count == 1 else f"{line_name} - Entry {entry_count}"
                
                # Create the production line
                production_line = ProductionLine.objects.create(
                    planning_board=board,
                    line_number=display_name,
                    # A Shift
                    a_shift_model=entry_data['a_shift']['model'],
                    a_shift_plan=entry_data['a_shift']['plan'],
                    a_shift_actual=entry_data['a_shift']['actual'],
                    a_shift_plan_change=entry_data['a_shift']['plan_change'],
                    a_shift_time=entry_data['a_shift']['time'],
                    a_shift_remarks=entry_data['a_shift']['remarks'],
                    # B Shift
                    b_shift_model=entry_data['b_shift']['model'],
                    b_shift_plan=entry_data['b_shift']['plan'],
                    b_shift_actual=entry_data['b_shift']['actual'],
                    b_shift_plan_change=entry_data['b_shift']['plan_change'],
                    b_shift_time=entry_data['b_shift']['time'],
                    b_shift_remarks=entry_data['b_shift']['remarks'],
                    # C Shift
                    c_shift_model=entry_data['c_shift']['model'],
                    c_shift_plan=entry_data['c_shift']['plan'],
                    c_shift_actual=entry_data['c_shift']['actual'],
                    c_shift_plan_change=entry_data['c_shift']['plan_change'],
                    c_shift_remarks=entry_data['c_shift']['remarks'],
                )
                
                print(f"Created production line: {display_name}")
                
    except Exception as e:
        print(f"Error creating production line entries for {line_name}: {e}")

def get_time_value(worksheet, row, col):
    """Get cell value as time, handling various time formats"""
    try:
        cell = worksheet.cell(row=row, column=col)
        value = cell.value
        
        if value is None:
            return None
            
        # If it's already a time object
        if hasattr(value, 'hour'):
            return value
            
        # Try to parse string time formats
        if isinstance(value, str):
            import re
            from datetime import time
            
            # Match HH:MM format
            time_match = re.search(r'(\d{1,2}):(\d{2})', value)
            if time_match:
                hour, minute = int(time_match.group(1)), int(time_match.group(2))
                if 0 <= hour <= 23 and 0 <= minute <= 59:
                    return time(hour, minute)
        
        return None
    except:
        return None

def extract_future_plans_fixed(worksheet, board):
    """Extract tomorrow and next day plans with manual column specification - FIXED"""
    try:
        print("=== EXTRACTING FUTURE PLANS FIXED ===")
        
        # MANUAL COLUMN SPECIFICATION - Adjust these based on your Excel layout
        
        # TOMORROW PLAN COLUMNS (typically starts around column T/U)
        tomorrow_config = {
            'model_col': 21,     # Column T - TOMORROW MODEL
            'a_shift_col': 22,   # Column U - TOMORROW A SHIFT  
            'b_shift_col': 23,   # Column V - TOMORROW B SHIFT
            'c_shift_col': 24,   # Column W - TOMORROW C SHIFT
            'remarks_col': 25,   # Column X - TOMORROW REMARKS
            'start_row': 7,      # Data starts at row 7
            'end_row': 46        # Data ends at row 30
        }
        
        # NEXT DAY PLAN COLUMNS (typically starts around column Y/Z)
        next_day_config = {
            'model_col': 26,     # Column Y - NEXT DAY MODEL
            'a_shift_col': 27,   # Column Z - NEXT DAY A SHIFT
            'b_shift_col': 28,   # Column AA - NEXT DAY B SHIFT
            'c_shift_col': 29,   # Column AB - NEXT DAY C SHIFT
            'remarks_col': 30,   # Column AC - NEXT DAY REMARKS
            'start_row': 7,      # Data starts at row 7
            'end_row': 46        # Data ends at row 30
        }
        
        print("Extracting TOMORROW plans with config:", tomorrow_config)
        extract_plan_with_config_fixed(worksheet, board, 'tomorrow', tomorrow_config)
        
        print("Extracting NEXT DAY plans with config:", next_day_config)
        extract_plan_with_config_fixed(worksheet, board, 'next_day', next_day_config)
        
    except Exception as e:
        print(f"Error extracting future plans: {e}")

def extract_plan_with_config_fixed(worksheet, board, plan_type, config):
    """Extract plan data using manual configuration - FIXED"""
    try:
        print(f"\n=== EXTRACTING {plan_type.upper()} PLANS ===")
        print(f"Config: {config}")
        
        created_count = 0
        
        for row in range(config['start_row'], config['end_row']):
            model = get_cell_value(worksheet, row, config['model_col'])
            
            if model and model.strip() and len(model.strip()) > 1:
                model = model.strip()
                
                # Skip headers and common non-data entries
                if model.upper() in ['MODEL', 'SHIFT', 'REMARKS', 'A', 'B', 'C', 'PLAN', 'ASSY', 'DAY', 'TOMORROW', 'NEXT']:
                    print(f"  Skipping header row {row}: '{model}'")
                    continue
                
                # Get shift data
                a_shift = get_numeric_value(worksheet, row, config['a_shift_col'])
                b_shift = get_numeric_value(worksheet, row, config['b_shift_col'])
                c_shift = get_numeric_value(worksheet, row, config['c_shift_col'])
                remarks = get_cell_value(worksheet, row, config['remarks_col'])
                
                # Debug print
                print(f"  {plan_type} Row {row}: Model='{model}', A={a_shift}, B={b_shift}, C={c_shift}, Remarks='{remarks}'")
                
                # Create database entry if we have model name (quantity can be 0)
                if plan_type == 'tomorrow':
                    TomorrowPlan.objects.create(
                        planning_board=board,
                        model=model,
                        a_shift=a_shift or 0,
                        b_shift=b_shift or 0,
                        c_shift=c_shift or 0,
                        remarks=remarks or ''
                    )
                else:  # next_day
                    NextDayPlan.objects.create(
                        planning_board=board,
                        model=model,
                        a_shift=a_shift or 0,
                        b_shift=b_shift or 0,
                        c_shift=c_shift or 0,
                        remarks=remarks or ''
                    )
                
                created_count += 1
                print(f"  ✓ Created {plan_type} plan: {model}")
        
        print(f"=== COMPLETED {plan_type.upper()}: Created {created_count} entries ===\n")
                
    except Exception as e:
        print(f"Error extracting {plan_type} with config: {e}")
        import traceback
        traceback.print_exc()

def extract_additional_sections(worksheet, board):
    """Extract additional sections like Critical Parts, AFM, SPD, etc. - ENHANCED"""
    try:
        print("Starting to extract additional sections...")
        
        # Look for section keywords and their positions
        critical_row = find_section_header(worksheet, "CRITICAL")
        fcin_row = find_section_header(worksheet, "FCIN")
        iu_row = find_section_header(worksheet, "I/U")
        msil_row = find_section_header(worksheet, "MSIL")
        hmsi_row = find_section_header(worksheet, "HMSI")
        iymp_row = find_section_header(worksheet, "IYM")
        hmcl_row = find_section_header(worksheet, "HMCL")
        other_row = find_section_header(worksheet, "OTHER")
        
        # Extract sections with proper data filtering
        if critical_row:
            extract_critical_parts_fixed(worksheet, board, critical_row)
        
        if fcin_row:
            extract_afm_plans_fixed(worksheet, board, fcin_row, "FCIN")
            
        if iu_row:
            extract_afm_plans_fixed(worksheet, board, iu_row, "IU")
        
        if msil_row:
            extract_spd_plans_fixed(worksheet, board, msil_row, "MSIL")
            
        if hmsi_row:
            extract_spd_plans_fixed(worksheet, board, hmsi_row, "HMSI")
            
        if iymp_row:
            extract_spd_plans_fixed(worksheet, board, iymp_row, "IYM")
            
        if hmcl_row:
            extract_spd_plans_fixed(worksheet, board, hmcl_row, "HMCL")
        
        if other_row:
            extract_other_information_fixed(worksheet, board, other_row)
        
    except Exception as e:
        print(f"Error extracting additional sections: {e}")
        import traceback
        traceback.print_exc()

def extract_critical_parts_fixed(worksheet, board, start_row):
    """Extract critical parts - ENHANCED to scan more rows"""
    try:
        print(f"Extracting critical parts from row {start_row}")
        
        # Start from start_row + 2 to skip header row, scan more rows
        for row in range(start_row + 2, start_row + 25):
            part_name = get_cell_value(worksheet, row, 2)    # Column B
            supplier = get_cell_value(worksheet, row, 3)     # Column C  
            plan_qty = get_numeric_value(worksheet, row, 4)  # Column D
            receiving_time_str = get_cell_value(worksheet, row, 5) # Column E
            remarks = get_cell_value(worksheet, row, 6)      # Column F
            
            # Skip if no part name or empty
            if not part_name or not part_name.strip():
                continue
                
            # Skip header rows explicitly
            if part_name.upper() in ['PART NAME', 'PART', 'NAME', 'SUPPLIER', 'QTY', 'QUANTITY']:
                print(f"Skipping critical parts header: {part_name}")
                continue
            
            # Only process real data rows
            if len(part_name.strip()) > 2:
                CriticalPartStatus.objects.create(
                    planning_board=board,
                    part_name=part_name.strip(),
                    supplier=supplier.strip() if supplier else '',
                    plan_qty=plan_qty or 0,
                    remarks=remarks.strip() if remarks else ''
                )
                print(f"Created critical part: {part_name}")
                
    except Exception as e:
        print(f"Error extracting critical parts: {e}")

def extract_afm_plans_fixed(worksheet, board, start_row, plan_type):
    """Extract AFM plans - ENHANCED to scan more rows"""
    try:
        print(f"Extracting AFM {plan_type} from row {start_row}")
        
        # Determine column positions based on plan type
        if plan_type == "FCIN":
            part_col = 7   # Column G
            num_col = 8    # Column H
            qty_col = 9    # Column I  
            rem_col = 10   # Column J
        else:  # I/U
            part_col = 11  # Column K
            num_col = 12   # Column L
            qty_col = 13   # Column M
            rem_col = 14   # Column N
        
        # Extract data starting from row after header, scan more rows
        for row in range(start_row + 3, start_row + 25):
            part_name = get_cell_value(worksheet, row, part_col)
            part_number = get_cell_value(worksheet, row, num_col)
            plan_qty = get_numeric_value(worksheet, row, qty_col)
            remarks = get_cell_value(worksheet, row, rem_col)
            
            # Skip if no part name
            if not part_name or not part_name.strip():
                continue
                
            # Skip header rows
            if part_name.upper() in ['PART NAME', 'PART', 'NAME', 'SUPPLIER']:
                print(f"Skipping AFM header: {part_name}")
                continue
            
            # Only process real data
            if len(part_name.strip()) > 2:
                AFMPlan.objects.create(
                    planning_board=board,
                    plan_type=plan_type,
                    part_name=part_name.strip(),
                    part_number=part_number.strip() if part_number else '',
                    plan_qty=plan_qty or 0,
                    remarks=remarks.strip() if remarks else ''
                )
                print(f"Created AFM {plan_type} plan: {part_name}")
                
    except Exception as e:
        print(f"Error extracting AFM {plan_type}: {e}")

def extract_spd_plans_fixed(worksheet, board, start_row, customer):
    """Extract SPD plans - ENHANCED to scan more rows"""
    try:
        print(f"Extracting SPD {customer} from row {start_row}")
        
        # Column positions for different customers
        customer_cols = {
            'MSIL': {'part': 15, 'num': 16, 'qty': 17, 'rem': 18},  # O, P, Q, R
            'HMSI': {'part': 19, 'num': 20, 'qty': 21},             # S, T, U
            'IYM':  {'part': 22, 'num': 23, 'qty': 24},             # V, W, X
            'HMCL': {'part': 25, 'num': 26, 'qty': 27}              # Y, Z, AA
        }
        
        cols = customer_cols.get(customer, customer_cols['MSIL'])
        
        # Extract data starting from row after header, scan more rows
        for row in range(start_row + 2, start_row + 30):
            part_name = get_cell_value(worksheet, row, cols['part'])
            part_number = get_cell_value(worksheet, row, cols['num'])
            plan_qty = get_numeric_value(worksheet, row, cols['qty'])
            
            # Remarks column only for customers that have it
            remarks_col = cols.get('rem')
            remarks = get_cell_value(worksheet, row, remarks_col) if remarks_col else ''
            
            # Skip if no part name
            if not part_name or not part_name.strip():
                continue
                
            # Skip header rows
            if part_name.upper() in ['PART NAME', 'PART', 'NAME', 'SUPPLIER']:
                print(f"Skipping SPD header: {part_name}")
                continue
            
            # Only process real data
            if len(part_name.strip()) > 2:
                SPDPlan.objects.create(
                    planning_board=board,
                    customer=customer,
                    part_name=part_name.strip(),
                    part_number=part_number.strip() if part_number else '',
                    plan_qty=plan_qty or 0,
                    remarks=remarks.strip() if remarks else ''
                )
                print(f"Created SPD {customer} plan: {part_name}")
                
    except Exception as e:
        print(f"Error extracting SPD {customer}: {e}")

def extract_other_information_fixed(worksheet, board, start_row):
    """Extract other information - ENHANCED to scan more rows"""
    try:
        print(f"Extracting other information from row {start_row}")
        
        # Extract data starting from row after header, scan more rows
        for row in range(start_row + 2, start_row + 25):
            part_name = get_cell_value(worksheet, row, 28)   # Column AA
            qty = get_numeric_value(worksheet, row, 29)      # Column AB
            target_date_str = get_cell_value(worksheet, row, 30) # Column AC
            remarks = get_cell_value(worksheet, row, 31)     # Column AD
            
            # Skip if no part name
            if not part_name or not part_name.strip():
                continue
                
            # Skip header rows
            if part_name.upper() in ['PART NAME', 'PART', 'NAME']:
                print(f"Skipping other info header: {part_name}")
                continue
            
            # Only process real data
            if len(part_name.strip()) > 2:
                # Use current date as default target date
                target_date = timezone.now().date()
                
                OtherInformation.objects.create(
                    planning_board=board,
                    part_name=part_name.strip(),
                    qty=qty or 0,
                    target_date=target_date,
                    remarks=remarks.strip() if remarks else ''
                )
                print(f"Created other information: {part_name}")
                
    except Exception as e:
        print(f"Error extracting other information: {e}")

def find_section_header(worksheet, keyword):
    """Find the row containing a specific section header keyword"""
    try:
        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell_value = get_cell_value(worksheet, row, col)
                if cell_value and keyword.upper() in cell_value.upper():
                    print(f"Found '{keyword}' at row {row}, col {col}: '{cell_value}'")
                    return row
        return None
    except Exception as e:
        print(f"Error finding section header for '{keyword}': {e}")
        return None

def get_cell_value(worksheet, row, col):
    """Get cell value as string, handling None values"""
    try:
        cell = worksheet.cell(row=row, column=col)
        value = cell.value
        if value is None:
            return ''
        return str(value).strip()
    except:
        return ''

def get_numeric_value(worksheet, row, col):
    """Get cell value as number, handling None and non-numeric values"""
    try:
        cell = worksheet.cell(row=row, column=col)
        value = cell.value
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return int(value) if value == int(value) else value
        # Try to convert string to number
        try:
            num_val = float(str(value).replace(',', ''))
            return int(num_val) if num_val == int(num_val) else num_val
        except:
            return None
    except:
        return None

@login_required
def export_to_excel(request, pk):
    """Export planning board data to Excel"""
    board = get_object_or_404(PlanningBoard, pk=pk, created_by=request.user)
    
    # Create new workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Planning Board"
    
    # Set up the Excel structure similar to the original format
    worksheet['B2'] = f"MEETING TIME: {board.meeting_time if board.meeting_time else ''}"
    worksheet['C2'] = board.title
    
    # Dates
    worksheet['C3'] = f"DATE:- {board.today_date}"
    worksheet['T3'] = f"DATE:- {board.tomorrow_date}"
    worksheet['Y3'] = f"DATE:- {board.next_day_date}"
    
    # Section headers
    worksheet['C4'] = "TODAY ASSY PLAN"
    worksheet['T4'] = "TOMORROW ASSY PLAN"
    worksheet['Y4'] = "NEXT DAY ASSY PLAN"
    
    # Column headers for production lines
    headers = ['LINE NO.', 'MODEL', 'PLAN', 'ACTUAL', 'PLAN CHANGE', 'TIME', 'REMARKS']
    for i, header in enumerate(headers, start=2):
        worksheet.cell(row=6, column=i, value=header)
    
    # Production line data
    row = 7
    for line in board.production_lines.all():
        worksheet.cell(row=row, column=2, value=line.line_number)
        # A Shift
        worksheet.cell(row=row, column=3, value=line.a_shift_model)
        worksheet.cell(row=row, column=4, value=line.a_shift_plan)
        worksheet.cell(row=row, column=5, value=line.a_shift_actual)
        worksheet.cell(row=row, column=6, value=line.a_shift_plan_change)
        worksheet.cell(row=row, column=7, value=str(line.a_shift_time) if line.a_shift_time else '')
        worksheet.cell(row=row, column=8, value=line.a_shift_remarks)
        row += 1
    
    # Tomorrow plans
    row = 7
    for plan in board.tomorrow_plans.all():
        worksheet.cell(row=row, column=20, value=plan.model)
        worksheet.cell(row=row, column=21, value=plan.a_shift)
        worksheet.cell(row=row, column=22, value=plan.b_shift)
        worksheet.cell(row=row, column=23, value=plan.c_shift)
        worksheet.cell(row=row, column=24, value=plan.remarks)
        row += 1
    
    # Next day plans
    row = 7
    for plan in board.next_day_plans.all():
        worksheet.cell(row=row, column=25, value=plan.model)
        worksheet.cell(row=row, column=26, value=plan.a_shift)
        worksheet.cell(row=row, column=27, value=plan.b_shift)
        worksheet.cell(row=row, column=28, value=plan.c_shift)
        worksheet.cell(row=row, column=29, value=plan.remarks)
        row += 1
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="planning_board_{board.today_date}.xlsx"'
    
    # Save workbook to response
    workbook.save(response)
    return response

@login_required
def ajax_add_production_line(request):
    """AJAX view to add new production line form"""
    if request.method == 'POST':
        board_id = request.POST.get('board_id')
        board = get_object_or_404(PlanningBoard, pk=board_id, created_by=request.user)
        
        # Get the current count to set the form index
        count = board.production_lines.count()
        
        # Create empty form
        formset = ProductionLineFormSet(instance=board)
        empty_form = formset.empty_form
        
        # Replace __prefix__ with actual index
        form_html = str(empty_form).replace('__prefix__', str(count))
        
        return JsonResponse({'form_html': form_html})
    
    return JsonResponse({'error': 'Invalid request'})

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Q
from datetime import datetime, timedelta
from django.contrib import messages
import logging

logger = logging.getLogger(__name__)

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Q
from datetime import datetime, timedelta
from django.contrib import messages
import logging

logger = logging.getLogger(__name__)

@login_required
def planning_board_dashboard(request):
    """Enhanced dashboard view with improved filtering capabilities"""
    
    # Get filter parameters
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    status_filter = request.GET.get('status', '').strip()
    
    # Start with all boards for the user
    boards_query = PlanningBoard.objects.filter(created_by=request.user)
    
    # Current date for calculations
    today = timezone.now().date()
    
    # Track if any filters were applied
    filters_applied = bool(date_from or date_to or status_filter)
    
    try:
        # Apply date filtering with better error handling
        if date_from:
            try:
                from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
                boards_query = boards_query.filter(today_date__gte=from_date)
                logger.debug(f"Applied date_from filter: {from_date}")
            except ValueError as e:
                logger.warning(f"Invalid date_from format: {date_from}")
                messages.warning(request, f"Invalid 'From Date' format: {date_from}")
                date_from = None  # Reset invalid date
        
        if date_to:
            try:
                to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
                boards_query = boards_query.filter(today_date__lte=to_date)
                logger.debug(f"Applied date_to filter: {to_date}")
            except ValueError as e:
                logger.warning(f"Invalid date_to format: {date_to}")
                messages.warning(request, f"Invalid 'To Date' format: {date_to}")
                date_to = None  # Reset invalid date
        
        # Apply status filtering with improved logic
        if status_filter:
            if status_filter == 'today':
                boards_query = boards_query.filter(today_date=today)
                # Auto-set date fields for consistency
                if not date_from:
                    date_from = today.strftime('%Y-%m-%d')
                if not date_to:
                    date_to = today.strftime('%Y-%m-%d')
                    
            elif status_filter == 'recent':
                three_days_ago = today - timedelta(days=3)
                boards_query = boards_query.filter(today_date__gte=three_days_ago)
                # Auto-set date fields for consistency
                if not date_from:
                    date_from = three_days_ago.strftime('%Y-%m-%d')
                if not date_to:
                    date_to = today.strftime('%Y-%m-%d')
                    
            elif status_filter == 'this_week':
                # Get Monday of current week
                week_start = today - timedelta(days=today.weekday())
                boards_query = boards_query.filter(today_date__gte=week_start)
                # Auto-set date fields for consistency
                if not date_from:
                    date_from = week_start.strftime('%Y-%m-%d')
                if not date_to:
                    date_to = today.strftime('%Y-%m-%d')
                    
            elif status_filter == 'this_month':
                month_start = today.replace(day=1)
                boards_query = boards_query.filter(today_date__gte=month_start)
                # Auto-set date fields for consistency
                if not date_from:
                    date_from = month_start.strftime('%Y-%m-%d')
                if not date_to:
                    date_to = today.strftime('%Y-%m-%d')
            
            logger.debug(f"Applied status filter: {status_filter}")
    
    except Exception as e:
        logger.error(f"Error applying filters: {e}")
        messages.error(request, "Error applying filters. Please try again.")
        # Reset to no filters
        date_from = date_to = status_filter = None
        boards_query = PlanningBoard.objects.filter(created_by=request.user)
    
    # If no filters applied and no request parameters, default to today's boards
    if not filters_applied and not request.GET:
        logger.debug("No filters applied, redirecting to today's filter")
        return redirect(f"{request.path}?date_from={today}&date_to={today}&status=today")
    
    # Order by most recent first and get the filtered results
    filtered_boards = boards_query.order_by('-created_at')
    
    # Get recent boards (last 5 regardless of filters for sidebar/stats)
    recent_boards = PlanningBoard.objects.filter(
        created_by=request.user
    ).order_by('-created_at')[:5]
    
    # Calculate comprehensive statistics
    user_boards = PlanningBoard.objects.filter(created_by=request.user)
    
    total_boards = user_boards.count()
    
    # Get total production lines for user's boards
    total_production_lines = ProductionLine.objects.filter(
        planning_board__created_by=request.user
    ).values('line_number').distinct().count()
    
    # Today's boards count
    today_boards_count = user_boards.filter(today_date=today).count()
    
    # This week's boards count
    week_start = today - timedelta(days=today.weekday())
    week_boards_count = user_boards.filter(today_date__gte=week_start).count()
    
    # This month's boards count
    month_start = today.replace(day=1)
    month_boards_count = user_boards.filter(today_date__gte=month_start).count()
    
    # Get total uploads if ExcelUpload model exists
    total_uploads = 0
    try:
        total_uploads = ExcelUpload.objects.filter(uploaded_by=request.user).count()
    except:
        # ExcelUpload model might not exist
        pass
    
    # Validate date range
    date_range_valid = True
    if date_from and date_to:
        try:
            from_date_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            to_date_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            if from_date_obj > to_date_obj:
                messages.warning(request, "'From Date' cannot be later than 'To Date'")
                date_range_valid = False
        except:
            pass
    
    # Prepare context
    context = {
        'filtered_boards': filtered_boards,
        'recent_boards': recent_boards,
        'total_boards': total_boards,
        'total_production_lines': total_production_lines,
        'total_uploads': total_uploads,
        'today_boards_count': today_boards_count,
        'week_boards_count': week_boards_count,
        'month_boards_count': month_boards_count,
        'filtered_count': filtered_boards.count(),
        
        # Filter values to maintain state
        'filter_date_from': date_from,
        'filter_date_to': date_to,
        'filter_status': status_filter,
        'filters_applied': filters_applied,
        'date_range_valid': date_range_valid,
        
        # Helper dates
        'today_date': today.strftime('%Y-%m-%d'),
        'week_start_date': week_start.strftime('%Y-%m-%d'),
        'month_start_date': month_start.strftime('%Y-%m-%d'),
        
        # Debug info (remove in production)
        'debug_info': {
            'query_count': filtered_boards.count(),
            'filters': {
                'date_from': date_from,
                'date_to': date_to,
                'status': status_filter,
            },
            'sql_query': str(filtered_boards.query) if filtered_boards else None,
        }
    }
    
    logger.debug(f"Dashboard context prepared with {filtered_boards.count()} filtered boards")
    
    return render(request, 'planning_board/dashboard.html', context)










# Add this to your views.py file
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_http_methods
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from datetime import datetime
from django.utils import timezone

@login_required
@require_http_methods(["POST"])
def inline_update_board(request, pk):
    """Handle inline updates for planning board and related data"""
    try:
        board = get_object_or_404(PlanningBoard, pk=pk, created_by=request.user)
        
        # Parse JSON data
        data = json.loads(request.body)
        print(f"Received data: {data}")  # Debug logging
        
        # Process board-level updates
        if 'board' in data and 'main' in data['board']:
            board_updates = data['board']['main']
            for field, value in board_updates.items():
                if hasattr(board, field):
                    # Handle different field types
                    if field in ['today_date', 'tomorrow_date', 'next_day_date']:
                        if value:
                            try:
                                setattr(board, field, datetime.strptime(value, '%Y-%m-%d').date())
                            except ValueError:
                                print(f"Invalid date format for {field}: {value}")
                                continue
                    elif field == 'meeting_time':
                        if value:
                            try:
                                setattr(board, field, datetime.strptime(value, '%H:%M').time())
                            except ValueError:
                                print(f"Invalid time format for {field}: {value}")
                                continue
                    else:
                        setattr(board, field, value or None)
            board.save()
            print(f"Updated board: {board.title}")
        
        # Process production line updates
        if 'production_line' in data:
            for line_id, updates in data['production_line'].items():
                if line_id.startswith('new_'):
                    # Create new production line
                    try:
                        ProductionLine.objects.create(
                            planning_board=board,
                            **process_production_line_data(updates)
                        )
                        print(f"Created new production line with temp ID: {line_id}")
                    except Exception as e:
                        print(f"Error creating production line: {e}")
                        continue
                else:
                    # Update existing production line
                    try:
                        line = ProductionLine.objects.get(id=line_id, planning_board=board)
                        for field, value in updates.items():
                            processed_value = process_field_value(field, value)
                            if hasattr(line, field):
                                setattr(line, field, processed_value)
                        line.save()
                        print(f"Updated production line ID: {line_id}")
                    except ProductionLine.DoesNotExist:
                        print(f"Production line not found: {line_id}")
                        continue
                    except Exception as e:
                        print(f"Error updating production line {line_id}: {e}")
                        continue
        
        # Process tomorrow plan updates
        if 'tomorrow_plan' in data:
            for plan_id, updates in data['tomorrow_plan'].items():
                if plan_id.startswith('new_'):
                    # Create new plan
                    try:
                        TomorrowPlan.objects.create(
                            planning_board=board,
                            **process_plan_data(updates)
                        )
                        print(f"Created new tomorrow plan with temp ID: {plan_id}")
                    except Exception as e:
                        print(f"Error creating tomorrow plan: {e}")
                        continue
                else:
                    # Update existing plan
                    try:
                        plan = TomorrowPlan.objects.get(id=plan_id, planning_board=board)
                        for field, value in updates.items():
                            processed_value = process_field_value(field, value)
                            if hasattr(plan, field):
                                setattr(plan, field, processed_value)
                        plan.save()
                        print(f"Updated tomorrow plan ID: {plan_id}")
                    except TomorrowPlan.DoesNotExist:
                        print(f"Tomorrow plan not found: {plan_id}")
                        continue
                    except Exception as e:
                        print(f"Error updating tomorrow plan {plan_id}: {e}")
                        continue
        
        # Process next day plan updates
        if 'next_day_plan' in data:
            for plan_id, updates in data['next_day_plan'].items():
                if plan_id.startswith('new_'):
                    # Create new plan
                    try:
                        NextDayPlan.objects.create(
                            planning_board=board,
                            **process_plan_data(updates)
                        )
                        print(f"Created new next day plan with temp ID: {plan_id}")
                    except Exception as e:
                        print(f"Error creating next day plan: {e}")
                        continue
                else:
                    # Update existing plan
                    try:
                        plan = NextDayPlan.objects.get(id=plan_id, planning_board=board)
                        for field, value in updates.items():
                            processed_value = process_field_value(field, value)
                            if hasattr(plan, field):
                                setattr(plan, field, processed_value)
                        plan.save()
                        print(f"Updated next day plan ID: {plan_id}")
                    except NextDayPlan.DoesNotExist:
                        print(f"Next day plan not found: {plan_id}")
                        continue
                    except Exception as e:
                        print(f"Error updating next day plan {plan_id}: {e}")
                        continue
        
        # Process critical part updates
        if 'critical_part' in data:
            for part_id, updates in data['critical_part'].items():
                if part_id.startswith('new_'):
                    # Create new critical part
                    try:
                        CriticalPartStatus.objects.create(
                            planning_board=board,
                            **process_critical_part_data(updates)
                        )
                        print(f"Created new critical part with temp ID: {part_id}")
                    except Exception as e:
                        print(f"Error creating critical part: {e}")
                        continue
                else:
                    # Update existing critical part
                    try:
                        part = CriticalPartStatus.objects.get(id=part_id, planning_board=board)
                        for field, value in updates.items():
                            processed_value = process_field_value(field, value)
                            if hasattr(part, field):
                                setattr(part, field, processed_value)
                        part.save()
                        print(f"Updated critical part ID: {part_id}")
                    except CriticalPartStatus.DoesNotExist:
                        print(f"Critical part not found: {part_id}")
                        continue
                    except Exception as e:
                        print(f"Error updating critical part {part_id}: {e}")
                        continue
        
        # Process AFM plan updates
        if 'afm_plan' in data:
            for plan_id, updates in data['afm_plan'].items():
                if plan_id.startswith('new_'):
                    # Create new AFM plan
                    try:
                        AFMPlan.objects.create(
                            planning_board=board,
                            **process_afm_plan_data(updates)
                        )
                        print(f"Created new AFM plan with temp ID: {plan_id}")
                    except Exception as e:
                        print(f"Error creating AFM plan: {e}")
                        continue
                else:
                    # Update existing AFM plan
                    try:
                        plan = AFMPlan.objects.get(id=plan_id, planning_board=board)
                        for field, value in updates.items():
                            processed_value = process_field_value(field, value)
                            if hasattr(plan, field):
                                setattr(plan, field, processed_value)
                        plan.save()
                        print(f"Updated AFM plan ID: {plan_id}")
                    except AFMPlan.DoesNotExist:
                        print(f"AFM plan not found: {plan_id}")
                        continue
                    except Exception as e:
                        print(f"Error updating AFM plan {plan_id}: {e}")
                        continue
        
        # Process SPD plan updates
        if 'spd_plan' in data:
            for plan_id, updates in data['spd_plan'].items():
                if plan_id.startswith('new_'):
                    # Create new SPD plan
                    try:
                        SPDPlan.objects.create(
                            planning_board=board,
                            **process_spd_plan_data(updates)
                        )
                        print(f"Created new SPD plan with temp ID: {plan_id}")
                    except Exception as e:
                        print(f"Error creating SPD plan: {e}")
                        continue
                else:
                    # Update existing SPD plan
                    try:
                        plan = SPDPlan.objects.get(id=plan_id, planning_board=board)
                        for field, value in updates.items():
                            processed_value = process_field_value(field, value)
                            if hasattr(plan, field):
                                setattr(plan, field, processed_value)
                        plan.save()
                        print(f"Updated SPD plan ID: {plan_id}")
                    except SPDPlan.DoesNotExist:
                        print(f"SPD plan not found: {plan_id}")
                        continue
                    except Exception as e:
                        print(f"Error updating SPD plan {plan_id}: {e}")
                        continue
        
        # Process other information updates
        if 'other_info' in data:
            for info_id, updates in data['other_info'].items():
                if info_id.startswith('new_'):
                    # Create new other info
                    try:
                        OtherInformation.objects.create(
                            planning_board=board,
                            **process_other_info_data(updates)
                        )
                        print(f"Created new other info with temp ID: {info_id}")
                    except Exception as e:
                        print(f"Error creating other info: {e}")
                        continue
                else:
                    # Update existing other info
                    try:
                        info = OtherInformation.objects.get(id=info_id, planning_board=board)
                        for field, value in updates.items():
                            processed_value = process_field_value(field, value)
                            if hasattr(info, field):
                                setattr(info, field, processed_value)
                        info.save()
                        print(f"Updated other info ID: {info_id}")
                    except OtherInformation.DoesNotExist:
                        print(f"Other info not found: {info_id}")
                        continue
                    except Exception as e:
                        print(f"Error updating other info {info_id}: {e}")
                        continue
        
        # Process deletions
        if 'delete' in data:
            for model_type, ids in data['delete'].items():
                for obj_id in ids:
                    try:
                        if model_type == 'production_line':
                            ProductionLine.objects.get(id=obj_id, planning_board=board).delete()
                            print(f"Deleted production line ID: {obj_id}")
                        elif model_type == 'tomorrow_plan':
                            TomorrowPlan.objects.get(id=obj_id, planning_board=board).delete()
                            print(f"Deleted tomorrow plan ID: {obj_id}")
                        elif model_type == 'next_day_plan':
                            NextDayPlan.objects.get(id=obj_id, planning_board=board).delete()
                            print(f"Deleted next day plan ID: {obj_id}")
                        elif model_type == 'critical_part':
                            CriticalPartStatus.objects.get(id=obj_id, planning_board=board).delete()
                            print(f"Deleted critical part ID: {obj_id}")
                        elif model_type == 'afm_plan':
                            AFMPlan.objects.get(id=obj_id, planning_board=board).delete()
                            print(f"Deleted AFM plan ID: {obj_id}")
                        elif model_type == 'spd_plan':
                            SPDPlan.objects.get(id=obj_id, planning_board=board).delete()
                            print(f"Deleted SPD plan ID: {obj_id}")
                        elif model_type == 'other_info':
                            OtherInformation.objects.get(id=obj_id, planning_board=board).delete()
                            print(f"Deleted other info ID: {obj_id}")
                    except Exception as e:
                        print(f"Error deleting {model_type} ID {obj_id}: {e}")
                        continue
        
        return JsonResponse({
            'success': True, 
            'message': 'Changes saved successfully',
            'board_id': board.pk
        })
        
    except json.JSONDecodeError as e:
        print(f"JSON decode error: {e}")
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        print(f"Unexpected error in inline_update_board: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

def process_field_value(field, value):
    """Process field value based on field type"""
    if not value or value == '-':
        return None
    
    # Handle time fields
    if 'time' in field and value:
        try:
            return datetime.strptime(value, '%H:%M').time()
        except ValueError:
            print(f"Invalid time format for {field}: {value}")
            return None
    
    # Handle date fields
    if 'date' in field and value:
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except ValueError:
            print(f"Invalid date format for {field}: {value}")
            return None
    
    # Handle datetime fields
    if field == 'receiving_time' and value:
        try:
            return datetime.strptime(value, '%Y-%m-%dT%H:%M')
        except ValueError:
            print(f"Invalid datetime format for {field}: {value}")
            return None
    
    # Handle numeric fields
    numeric_fields = [
        'plan', 'actual', 'plan_change', 'plan_qty', 'qty',
        'a_shift_plan', 'a_shift_actual', 'a_shift_plan_change',
        'b_shift_plan', 'b_shift_actual', 'b_shift_plan_change',
        'c_shift_plan', 'c_shift_actual', 'c_shift_plan_change',
        'a_shift', 'b_shift', 'c_shift'
    ]
    
    if field in numeric_fields and value:
        try:
            # Try integer first, then float
            if '.' in str(value):
                return float(value)
            else:
                return int(value)
        except (ValueError, TypeError):
            print(f"Invalid numeric value for {field}: {value}")
            return 0
    
    # Return string value for everything else
    return str(value).strip() if value else ''

def process_production_line_data(data):
    """Process production line data for creation"""
    processed = {}
    for field, value in data.items():
        processed[field] = process_field_value(field, value)
    
    # Ensure required fields have default values
    if 'line_number' not in processed or not processed['line_number']:
        processed['line_number'] = 'New Line'
    
    return processed

def process_plan_data(data):
    """Process plan data for creation (tomorrow/next day plans)"""
    processed = {}
    for field, value in data.items():
        if field in ['a_shift', 'b_shift', 'c_shift'] and value:
            try:
                processed[field] = int(value) if str(value).isdigit() else float(value)
            except (ValueError, TypeError):
                processed[field] = 0
        else:
            processed[field] = str(value).strip() if value else ''
    
    # Ensure required fields have default values
    if 'model' not in processed or not processed['model']:
        processed['model'] = 'New Model'
    
    return processed

def process_critical_part_data(data):
    """Process critical part data for creation"""
    processed = {}
    for field, value in data.items():
        if field == 'plan_qty' and value:
            try:
                processed[field] = int(value) if str(value).isdigit() else float(value)
            except (ValueError, TypeError):
                processed[field] = 0
        elif field == 'receiving_time' and value:
            try:
                processed[field] = datetime.strptime(value, '%Y-%m-%dT%H:%M')
            except ValueError:
                processed[field] = None
        else:
            processed[field] = str(value).strip() if value else ''
    
    # Ensure required fields have default values
    if 'part_name' not in processed or not processed['part_name']:
        processed['part_name'] = 'New Critical Part'
    
    return processed

def process_afm_plan_data(data):
    """Process AFM plan data for creation"""
    processed = {}
    for field, value in data.items():
        if field == 'plan_qty' and value:
            try:
                processed[field] = int(value) if str(value).isdigit() else float(value)
            except (ValueError, TypeError):
                processed[field] = 0
        else:
            processed[field] = str(value).strip() if value else ''
    
    # Ensure required fields have default values
    if 'part_name' not in processed or not processed['part_name']:
        processed['part_name'] = 'New AFM Part'
    if 'plan_type' not in processed or not processed['plan_type']:
        processed['plan_type'] = 'FCIN'
    
    return processed

def process_spd_plan_data(data):
    """Process SPD plan data for creation"""
    processed = {}
    for field, value in data.items():
        if field == 'plan_qty' and value:
            try:
                processed[field] = int(value) if str(value).isdigit() else float(value)
            except (ValueError, TypeError):
                processed[field] = 0
        else:
            processed[field] = str(value).strip() if value else ''
    
    # Ensure required fields have default values
    if 'part_name' not in processed or not processed['part_name']:
        processed['part_name'] = 'New SPD Part'
    if 'customer' not in processed or not processed['customer']:
        processed['customer'] = 'MSIL'
    
    return processed

def process_other_info_data(data):
    """Process other information data for creation"""
    processed = {}
    for field, value in data.items():
        if field == 'qty' and value:
            try:
                processed[field] = int(value) if str(value).isdigit() else float(value)
            except (ValueError, TypeError):
                processed[field] = 0
        elif field == 'target_date' and value:
            try:
                processed[field] = datetime.strptime(value, '%Y-%m-%d').date()
            except ValueError:
                processed[field] = timezone.now().date()
        else:
            processed[field] = str(value).strip() if value else ''
    
    # Ensure required fields have default values
    if 'part_name' not in processed or not processed['part_name']:
        processed['part_name'] = 'New Information'
    if 'target_date' not in processed:
        processed['target_date'] = timezone.now().date()
    
    return processed

    













#  new one 


import json
import time
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, StreamingHttpResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.cache import never_cache
from django.utils import timezone
from datetime import datetime
from django.db.models import Count, Q
from .models import (
    PlanningBoard, ProductionLine, TomorrowPlan, NextDayPlan,
    CriticalPartStatus, AFMPlan, SPDPlan, OtherInformation
)

@login_required
def live_view_page(request):
    """Render the live view page"""
    # Get user's planning boards for the dropdown
    boards = PlanningBoard.objects.filter(created_by=request.user).order_by('-created_at')[:10]
    
    context = {
        'boards': boards,
        'current_time': timezone.now(),
    }
    
    return render(request, 'planning_board/live_view.html', context)

@login_required
@never_cache
def get_board_sections_summary(request, board_id):
    """Get summary data for all sections of a planning board"""
    board = get_object_or_404(PlanningBoard, pk=board_id, created_by=request.user)
    
    # Get counts and last updated times for each section
    sections_data = {}
    
    # Today Assembly Plan (Production Lines)
    production_lines = board.production_lines.all()
    sections_data['today_assembly'] = {
        'count': production_lines.count(),
        'last_updated': board.updated_at,
        'status': 'active' if production_lines.exists() else 'empty'
    }
    
    # Tomorrow Assembly Plan
    tomorrow_plans = board.tomorrow_plans.all()
    sections_data['tomorrow_assembly'] = {
        'count': tomorrow_plans.count(),
        'last_updated': board.updated_at,
        'status': 'active' if tomorrow_plans.exists() else 'empty'
    }
    
    # Next Day Assembly Plan
    next_day_plans = board.next_day_plans.all()
    sections_data['next_day_assembly'] = {
        'count': next_day_plans.count(),
        'last_updated': board.updated_at,
        'status': 'active' if next_day_plans.exists() else 'empty'
    }
    
    # Critical Parts
    critical_parts = board.critical_parts.all()
    sections_data['critical_parts'] = {
        'count': critical_parts.count(),
        'last_updated': board.updated_at,
        'status': 'active' if critical_parts.exists() else 'empty'
    }
    
    # AFM Plans
    afm_plans = board.afm_plans.all()
    sections_data['afm_plans'] = {
        'count': afm_plans.count(),
        'last_updated': board.updated_at,
        'status': 'active' if afm_plans.exists() else 'empty'
    }
    
    # SPD Plans
    spd_plans = board.spd_plans.all()
    sections_data['spd_plans'] = {
        'count': spd_plans.count(),
        'last_updated': board.updated_at,
        'status': 'active' if spd_plans.exists() else 'empty'
    }
    
    # Other Information
    other_info = board.other_info.all()
    sections_data['other_info'] = {
        'count': other_info.count(),
        'last_updated': board.updated_at,
        'status': 'active' if other_info.exists() else 'empty'
    }
    
    return JsonResponse({
        'board_id': board.pk,
        'board_title': board.title,
        'board_date': board.today_date.strftime('%Y-%m-%d'),
        'sections': sections_data,
        'timestamp': timezone.now().isoformat()
    })

@login_required
@never_cache
def get_section_data(request, board_id, section):
    """Get detailed data for a specific section"""
    board = get_object_or_404(PlanningBoard, pk=board_id, created_by=request.user)
    
    data = {
        'section': section,
        'board_id': board.pk,
        'timestamp': timezone.now().isoformat(),
        'data': []
    }
    
    if section == 'today_assembly':
        production_lines = board.production_lines.all().order_by('line_number')
        data['title'] = 'Today Assembly Plan  (A SHIFT TIMING  6:00 AM TO 2:30 PM) --- (B SHIFT TIMING 2:30 PM  TO 11:00 PM) --- (C SHIFT TIMING SHIFT TIMING 11:00PM TO 6:00 AM)'
        data['headers'] = [
            'Line No.', 'A Shift Model', 'A Plan', 'A Actual', 'A Change', 'A Time',
            'B Shift Model', 'B Plan', 'B Actual', 'B Change', 'B Time',
            'C Shift Model', 'C Plan', 'C Actual', 'C Change', 'C Time'
        ]
        
        for line in production_lines:
            data['data'].append([
                line.line_number or '',
                line.a_shift_model or '',
                line.a_shift_plan or 0,
                line.a_shift_actual or 0,
                line.a_shift_plan_change or 0,
                str(line.a_shift_time) if line.a_shift_time else '6:00 AM TO 2:30 PM',
                line.b_shift_model or '',
                line.b_shift_plan or 0,
                line.b_shift_actual or 0,
                line.b_shift_plan_change or 0,
                str(line.b_shift_time) if line.b_shift_time else '2:30 PM  TO 11:00 PM',
                line.c_shift_model or '',
                line.c_shift_plan or 0,
                line.c_shift_actual or 0,
                line.c_shift_plan_change or 0,
                ( '11:00PM TO 6:00 AM')[:100]
            ])
    
    elif section == 'tomorrow_assembly':
        tomorrow_plans = board.tomorrow_plans.all().order_by('model')
        data['title'] = 'Tomorrow Assembly Plan'
        data['headers'] = ['Model', 'A Shift', 'B Shift', 'C Shift', 'Total', 'Remarks']
        
        for plan in tomorrow_plans:
            total = (plan.a_shift or 0) + (plan.b_shift or 0) + (plan.c_shift or 0)
            data['data'].append([
                plan.model or '',
                plan.a_shift or 0,
                plan.b_shift or 0,
                plan.c_shift or 0,
                total,
                (plan.remarks or '')[:100]
            ])
    
    elif section == 'next_day_assembly':
        next_day_plans = board.next_day_plans.all().order_by('model')
        data['title'] = 'Next Day Assembly Plan'
        data['headers'] = ['Model', 'A Shift', 'B Shift', 'C Shift', 'Total', 'Remarks']
        
        for plan in next_day_plans:
            total = (plan.a_shift or 0) + (plan.b_shift or 0) + (plan.c_shift or 0)
            data['data'].append([
                plan.model or '',
                plan.a_shift or 0,
                plan.b_shift or 0,
                plan.c_shift or 0,
                total,
                (plan.remarks or '')[:100]
            ])
    
    elif section == 'critical_parts':
        critical_parts = board.critical_parts.all().order_by('part_name')
        data['title'] = 'Critical Part Status'
        data['headers'] = ['Part Name', 'Supplier', 'Plan Qty', 'Receiving Time', 'Status', 'Remarks']
        
        for part in critical_parts:
            # Determine status based on receiving time
            status = 'Pending'
            if part.receiving_time:
                if part.receiving_time <= timezone.now():
                    status = 'Received'
                else:
                    status = 'Scheduled'
            
            data['data'].append([
                part.part_name or '',
                part.supplier or '',
                part.plan_qty or 0,
                part.receiving_time.strftime('%Y-%m-%d %H:%M') if part.receiving_time else '',
                status,
                (part.remarks or '')[:100]
            ])
    
    elif section == 'afm_plans':
        afm_plans = board.afm_plans.all().order_by('plan_type', 'part_name')
        data['title'] = 'AFM Plans'
        data['headers'] = ['Type', 'Part Name', 'Part Number', 'Plan Qty', 'Remarks']
        
        for plan in afm_plans:
            data['data'].append([
                plan.plan_type or '',
                plan.part_name or '',
                plan.part_number or '',
                plan.plan_qty or 0,
                (plan.remarks or '')[:100]
            ])
    
    elif section == 'spd_plans':
        spd_plans = board.spd_plans.all().order_by('customer', 'part_name')
        data['title'] = 'SPD Plans (Customer-wise)'
        data['headers'] = ['Customer', 'Part Name', 'Part Number', 'Plan Qty', 'Remarks']
        
        for plan in spd_plans:
            data['data'].append([
                plan.customer or '',
                plan.part_name or '',
                plan.part_number or '',
                plan.plan_qty or 0,
                (plan.remarks or '')[:100]
            ])
    
    elif section == 'other_info':
        other_info = board.other_info.all().order_by('target_date', 'part_name')
        data['title'] = 'Other Information'
        data['headers'] = ['Part Name', 'Quantity', 'Target Date', 'Remarks']
        
        for info in other_info:
            data['data'].append([
                info.part_name or '',
                info.qty or 0,
                info.target_date.strftime('%Y-%m-%d') if info.target_date else '',
                (info.remarks or '')[:100]
            ])
    
    return JsonResponse(data)

@login_required
@never_cache
def live_stream_section(request, board_id, section):
    """Server-Sent Events stream for real-time updates"""
    
    def event_stream():
        """Generator function for SSE stream"""
        board = get_object_or_404(PlanningBoard, pk=board_id, created_by=request.user)
        last_update = timezone.now()
        
        while True:
            try:
                # Check if board has been updated
                current_board = PlanningBoard.objects.get(pk=board_id, created_by=request.user)
                
                if current_board.updated_at > last_update:
                    # Get fresh data
                    response = get_section_data(request, board_id, section)
                    data = json.loads(response.content.decode('utf-8'))
                    
                    # Send SSE event
                    yield f"data: {json.dumps(data)}\n\n"
                    
                    last_update = current_board.updated_at
                
                # Send heartbeat every 30 seconds
                yield f"event: heartbeat\ndata: {json.dumps({'timestamp': timezone.now().isoformat()})}\n\n"
                
                time.sleep(5)  # Check for updates every 5 seconds
                
            except PlanningBoard.DoesNotExist:
                yield f"event: error\ndata: {json.dumps({'error': 'Board not found'})}\n\n"
                break
            except Exception as e:
                yield f"event: error\ndata: {json.dumps({'error': str(e)})}\n\n"
                time.sleep(10)  # Wait longer on error
    
    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['Access-Control-Allow-Origin'] = '*'
    response['Access-Control-Allow-Headers'] = 'Cache-Control'
    
    return response

@login_required
def get_user_planning_boards(request):
    """Get list of planning boards for the current user"""
    boards = PlanningBoard.objects.filter(created_by=request.user).order_by('-created_at')[:20]
    
    boards_data = []
    for board in boards:
        boards_data.append({
            'id': board.pk,
            'title': board.title,
            'date': board.today_date.strftime('%Y-%m-%d'),
            'created_at': board.created_at.strftime('%Y-%m-%d %H:%M'),
            'meeting_time': str(board.meeting_time) if board.meeting_time else None
        })
    
    return JsonResponse({
        'boards': boards_data,
        'total_count': boards.count()
    })

@login_required
@csrf_exempt
def trigger_board_update(request, board_id):
    """Trigger a manual update of the board's timestamp (for testing real-time updates)"""
    if request.method == 'POST':
        board = get_object_or_404(PlanningBoard, pk=board_id, created_by=request.user)
        board.updated_at = timezone.now()
        board.save(update_fields=['updated_at'])
        
        return JsonResponse({
            'success': True,
            'message': 'Board updated successfully',
            'updated_at': board.updated_at.isoformat()
        })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


#  new for rendering the new page


 


# Add these views to your existing views.py file

@login_required
def fullscreen_display(request, board_id, section):
    """Fullscreen industry-style display for selected section"""
    board = get_object_or_404(PlanningBoard, pk=board_id, created_by=request.user)
    
    # Validate section
    valid_sections = [
        'today_assembly', 'tomorrow_assembly', 'next_day_assembly',
        'critical_parts', 'afm_plans', 'spd_plans', 'other_info'
    ]
    
    if section not in valid_sections:
        messages.error(request, 'Invalid section selected')
        return redirect('planning_board:live_view')
    
    # Get section display configuration
    section_config = get_section_display_config(section)
    
    context = {
        'board': board,
        'section': section,
        'section_config': section_config,
        'current_time': timezone.now(),
        'refresh_url': f'/planning-board/api/board/{board_id}/section/{section}/',
        'stream_url': f'/planning-board/api/board/{board_id}/section/{section}/stream/',
    }
    
    return render(request, 'planning_board/fullscreen_display.html', context)

def get_section_display_config(section):
    """Get display configuration for each section"""
    configs = {
        'today_assembly': {
            'title': 'Today Assembly Plan',
            'subtitle': 'Real-time Production Status',
            'icon': '🏭',
            'color_scheme': 'blue',
            'layout': 'production_grid',
            'refresh_interval': 5000,
            'show_progress': True,
            'show_alerts': True,
        },
        'tomorrow_assembly': {
            'title': 'Tomorrow Assembly Plan',
            'subtitle': 'Next Day Production Schedule',
            'icon': '📅',
            'color_scheme': 'green',
            'layout': 'planning_grid',
            'refresh_interval': 10000,
            'show_progress': False,
            'show_alerts': False,
        },
        'next_day_assembly': {
            'title': 'Next Day Assembly Plan',
            'subtitle': 'Future Production Schedule',
            'icon': '⏭️',
            'color_scheme': 'purple',
            'layout': 'planning_grid',
            'refresh_interval': 15000,
            'show_progress': False,
            'show_alerts': False,
        },
        'critical_parts': {
            'title': 'Critical Parts Status',
            'subtitle': 'Supply Chain Monitoring',
            'icon': '⚠️',
            'color_scheme': 'red',
            'layout': 'status_grid',
            'refresh_interval': 3000,
            'show_progress': True,
            'show_alerts': True,
        },
        'afm_plans': {
            'title': 'AFM Production Plans',
            'subtitle': 'FCIN & I/U Operations',
            'icon': '🔧',
            'color_scheme': 'orange',
            'layout': 'operation_grid',
            'refresh_interval': 8000,
            'show_progress': True,
            'show_alerts': False,
        },
        'spd_plans': {
            'title': 'SPD Customer Plans',
            'subtitle': 'Customer-wise Production',
            'icon': '🏢',
            'color_scheme': 'teal',
            'layout': 'customer_grid',
            'refresh_interval': 10000,
            'show_progress': True,
            'show_alerts': False,
        },
        'other_info': {
            'title': 'Additional Information',
            'subtitle': 'Support & Maintenance',
            'icon': '📝',
            'color_scheme': 'gray',
            'layout': 'info_grid',
            'refresh_interval': 12000,
            'show_progress': False,
            'show_alerts': False,
        }
    }
    
    return configs.get(section, configs['today_assembly'])

@never_cache
def fullscreen_data_stream(request, board_id, section):
    """Enhanced streaming endpoint for fullscreen display with additional metadata"""
    
    def event_stream():
        board = get_object_or_404(PlanningBoard, pk=board_id, created_by=request.user)
        last_update = timezone.now()
        
        while True:
            try:
                # Check if board has been updated
                current_board = PlanningBoard.objects.get(pk=board_id, created_by=request.user)
                
                if current_board.updated_at > last_update:
                    # Get enhanced data with statistics
                    data = get_enhanced_section_data(board_id, section, request.user)
                    
                    # Send SSE event
                    yield f"data: {json.dumps(data)}\n\n"
                    
                    last_update = current_board.updated_at
                
                # Send heartbeat with system status
                system_status = {
                    'timestamp': timezone.now().isoformat(),
                    'board_id': board_id,
                    'section': section,
                    'connection_count': 1,  # Could track multiple connections
                    'server_time': timezone.now().strftime('%H:%M:%S'),
                    'server_date': timezone.now().strftime('%Y-%m-%d'),
                }
                
                yield f"event: heartbeat\ndata: {json.dumps(system_status)}\n\n"
                
                time.sleep(3)  # More frequent updates for fullscreen display
                
            except PlanningBoard.DoesNotExist:
                yield f"event: error\ndata: {json.dumps({'error': 'Board not found'})}\n\n"
                break
            except Exception as e:
                yield f"event: error\ndata: {json.dumps({'error': str(e)})}\n\n"
                time.sleep(5)
    
    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['Connection'] = 'keep-alive'
    response['Access-Control-Allow-Origin'] = '*'
    
    return response

def get_enhanced_section_data(board_id, section, user):
    """Get section data with additional statistics and metadata for fullscreen display"""
    board = PlanningBoard.objects.get(pk=board_id, created_by=user)
    
    base_data = {
        'section': section,
        'board_id': board.pk,
        'board_title': board.title,
        'board_date': board.today_date.strftime('%Y-%m-%d'),
        'timestamp': timezone.now().isoformat(),
        'last_updated': board.updated_at.isoformat(),
        'data': [],
        'statistics': {},
        'alerts': [],
        'status': 'active'
    }
    
    if section == 'today_assembly':
        production_lines = board.production_lines.all().order_by('line_number')
        base_data['title'] = 'Today Assembly Plan'
        base_data['headers'] = [
            'Line', 'A Shift Model', 'A Plan', 'A Actual', 'A %',
            'B Shift Model', 'B Plan', 'B Actual', 'B %',
            'C Shift Model', 'C Plan', 'C Actual', 'C %', 'Status'
        ]
        
        total_plan = 0
        total_actual = 0
        line_statuses = {'on_target': 0, 'behind': 0, 'ahead': 0}
        
        for line in production_lines:
            # Calculate efficiency percentages
            a_efficiency = calculate_efficiency(line.a_shift_plan, line.a_shift_actual)
            b_efficiency = calculate_efficiency(line.b_shift_plan, line.b_shift_actual)
            c_efficiency = calculate_efficiency(line.c_shift_plan, line.c_shift_actual)
            
            # Determine line status
            avg_efficiency = (a_efficiency + b_efficiency + c_efficiency) / 3 if any([a_efficiency, b_efficiency, c_efficiency]) else 0
            if avg_efficiency >= 95:
                status = 'On Target'
                line_statuses['on_target'] += 1
            elif avg_efficiency >= 85:
                status = 'Behind'
                line_statuses['behind'] += 1
            else:
                status = 'Critical'
                line_statuses['behind'] += 1
                
            # Add alert for critical lines
            if avg_efficiency < 85 and avg_efficiency > 0:
                base_data['alerts'].append({
                    'type': 'warning',
                    'message': f'{line.line_number}: Production below 85%',
                    'efficiency': avg_efficiency
                })
            
            total_plan += (line.a_shift_plan or 0) + (line.b_shift_plan or 0) + (line.c_shift_plan or 0)
            total_actual += (line.a_shift_actual or 0) + (line.b_shift_actual or 0) + (line.c_shift_actual or 0)
            
            base_data['data'].append([
                line.line_number or '',
                line.a_shift_model or '-',
                line.a_shift_plan or 0,
                line.a_shift_actual or 0,
                f"{a_efficiency}%",
                line.b_shift_model or '-',
                line.b_shift_plan or 0,
                line.b_shift_actual or 0,
                f"{b_efficiency}%",
                line.c_shift_model or '-',
                line.c_shift_plan or 0,
                line.c_shift_actual or 0,
                f"{c_efficiency}%",
                status
            ])
        
        base_data['statistics'] = {
            'total_lines': production_lines.count(),
            'total_plan': total_plan,
            'total_actual': total_actual,
            'overall_efficiency': calculate_efficiency(total_plan, total_actual),
            'lines_on_target': line_statuses['on_target'],
            'lines_behind': line_statuses['behind'],
            'lines_ahead': line_statuses['ahead'],
        }
    
    elif section == 'critical_parts':
        critical_parts = board.critical_parts.all().order_by('part_name')
        base_data['title'] = 'Critical Parts Status'
        base_data['headers'] = ['Part Name', 'Supplier', 'Plan Qty', 'Status', 'ETA', 'Risk Level']
        
        status_counts = {'received': 0, 'pending': 0, 'delayed': 0}
        risk_counts = {'low': 0, 'medium': 0, 'high': 0}
        
        for part in critical_parts:
            # Determine status and risk
            if part.receiving_time:
                if part.receiving_time <= timezone.now():
                    status = 'Received'
                    status_counts['received'] += 1
                    risk = 'Low'
                    risk_counts['low'] += 1
                elif part.receiving_time <= timezone.now() + timedelta(hours=2):
                    status = 'Due Soon'
                    status_counts['pending'] += 1
                    risk = 'Medium'
                    risk_counts['medium'] += 1
                else:
                    status = 'Scheduled'
                    status_counts['pending'] += 1
                    risk = 'Low'
                    risk_counts['low'] += 1
                    
                eta = part.receiving_time.strftime('%H:%M')
            else:
                status = 'TBD'
                eta = 'TBD'
                risk = 'High'
                status_counts['delayed'] += 1
                risk_counts['high'] += 1
                
                # Add alert for TBD parts
                base_data['alerts'].append({
                    'type': 'error',
                    'message': f'{part.part_name}: No receiving time set',
                    'supplier': part.supplier
                })
            
            base_data['data'].append([
                part.part_name or '',
                part.supplier or '',
                part.plan_qty or 0,
                status,
                eta,
                risk
            ])
        
        base_data['statistics'] = {
            'total_parts': critical_parts.count(),
            'received': status_counts['received'],
            'pending': status_counts['pending'],
            'delayed': status_counts['delayed'],
            'high_risk': risk_counts['high'],
            'medium_risk': risk_counts['medium'],
            'low_risk': risk_counts['low'],
        }
    
    # Add similar enhanced processing for other sections
    elif section in ['tomorrow_assembly', 'next_day_assembly']:
        if section == 'tomorrow_assembly':
            plans = board.tomorrow_plans.all().order_by('model')
            base_data['title'] = 'Tomorrow Assembly Plan'
        else:
            plans = board.next_day_plans.all().order_by('model')
            base_data['title'] = 'Next Day Assembly Plan'
            
        base_data['headers'] = ['Model', 'A Shift', 'B Shift', 'C Shift', 'Total', 'Priority']
        
        total_planned = 0
        shift_totals = {'a': 0, 'b': 0, 'c': 0}
        
        for plan in plans:
            a_shift = plan.a_shift or 0
            b_shift = plan.b_shift or 0
            c_shift = plan.c_shift or 0
            total = a_shift + b_shift + c_shift
            
            # Determine priority based on total quantity
            if total > 500:
                priority = 'High'
            elif total > 200:
                priority = 'Medium'
            else:
                priority = 'Low'
            
            shift_totals['a'] += a_shift
            shift_totals['b'] += b_shift
            shift_totals['c'] += c_shift
            total_planned += total
            
            base_data['data'].append([
                plan.model or '',
                a_shift,
                b_shift,
                c_shift,
                total,
                priority
            ])
        
        base_data['statistics'] = {
            'total_models': plans.count(),
            'total_planned': total_planned,
            'a_shift_total': shift_totals['a'],
            'b_shift_total': shift_totals['b'],
            'c_shift_total': shift_totals['c'],
        }
    
    elif section == 'afm_plans':
        afm_plans = board.afm_plans.all().order_by('plan_type', 'part_name')
        base_data['title'] = 'AFM Plans'
        base_data['headers'] = ['Type', 'Part Name', 'Part Number', 'Plan Qty', 'Remarks']
        
        type_counts = {'FCIN': 0, 'IU': 0}
        total_qty = 0
        
        for plan in afm_plans:
            type_counts[plan.plan_type] = type_counts.get(plan.plan_type, 0) + 1
            total_qty += plan.plan_qty or 0
            
            base_data['data'].append([
                plan.plan_type or '',
                plan.part_name or '',
                plan.part_number or '',
                plan.plan_qty or 0,
                (plan.remarks or '')[:50]
            ])
        
        base_data['statistics'] = {
            'total_plans': afm_plans.count(),
            'fcin_count': type_counts.get('FCIN', 0),
            'iu_count': type_counts.get('IU', 0),
            'total_quantity': total_qty,
        }
    
    elif section == 'spd_plans':
        spd_plans = board.spd_plans.all().order_by('customer', 'part_name')
        base_data['title'] = 'SPD Plans (Customer-wise)'
        base_data['headers'] = ['Customer', 'Part Name', 'Part Number', 'Plan Qty', 'Remarks']
        
        customer_counts = {}
        total_qty = 0
        
        for plan in spd_plans:
            customer_counts[plan.customer] = customer_counts.get(plan.customer, 0) + 1
            total_qty += plan.plan_qty or 0
            
            base_data['data'].append([
                plan.customer or '',
                plan.part_name or '',
                plan.part_number or '',
                plan.plan_qty or 0,
                (plan.remarks or '')[:50]
            ])
        
        base_data['statistics'] = {
            'total_plans': spd_plans.count(),
            'customer_breakdown': customer_counts,
            'total_quantity': total_qty,
        }
    
    elif section == 'other_info':
        other_info = board.other_info.all().order_by('target_date', 'part_name')
        base_data['title'] = 'Other Information'
        base_data['headers'] = ['Part Name', 'Quantity', 'Target Date', 'Days Left', 'Remarks']
        
        total_items = 0
        due_soon = 0
        
        for info in other_info:
            days_left = (info.target_date - timezone.now().date()).days if info.target_date else 0
            if days_left <= 3:
                due_soon += 1
            total_items += 1
            
            base_data['data'].append([
                info.part_name or '',
                info.qty or 0,
                info.target_date.strftime('%Y-%m-%d') if info.target_date else '',
                days_left,
                (info.remarks or '')[:50]
            ])
        
        base_data['statistics'] = {
            'total_items': total_items,
            'due_soon': due_soon,
        }
    
    return base_data

def calculate_efficiency(plan, actual):
    """Calculate efficiency percentage"""
    if not plan or plan == 0:
        return 0
    return round((actual or 0) / plan * 100, 1)