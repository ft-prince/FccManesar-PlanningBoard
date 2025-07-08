# management/commands/create_mock_excel.py
# Create this file in: your_app/management/commands/create_mock_excel.py

import openpyxl
from django.core.management.base import BaseCommand
from django.conf import settings
import os
from datetime import datetime, time

class Command(BaseCommand):
    help = 'Create a mock Excel file with comprehensive planning board data'

    def add_arguments(self, parser):
        parser.add_argument(
            '--filename',
            type=str,
            default='mock_planning_board.xlsx',
            help='Name of the Excel file to create'
        )

    def handle(self, *args, **options):
        filename = options['filename']
        
        # Create workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Planning Board"
        
        self.stdout.write("Creating mock Excel file...")
        
        # Populate all sections
        self.populate_headers(worksheet)
        self.populate_production_lines(worksheet)
        self.populate_tomorrow_plans(worksheet)
        self.populate_next_day_plans(worksheet)
        self.populate_section_headers(worksheet)
        self.populate_critical_parts(worksheet)
        self.populate_afm_plans(worksheet)
        self.populate_spd_plans(worksheet)
        self.populate_other_information(worksheet)
        
        # Save file
        file_path = os.path.join(settings.MEDIA_ROOT, 'uploads', 'excel', filename)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        workbook.save(file_path)
        
        self.stdout.write(
            self.style.SUCCESS(f'Successfully created mock Excel file: {file_path}')
        )

    def populate_headers(self, worksheet):
        """Populate header information"""
        worksheet['B2'] = "MEETING TIME: 09:30"
        worksheet['C2'] = "PRODUCTION PLANNING CONTROL DISPLAY BOARD"
        worksheet['C3'] = "DATE:- 2025-07-08"
        worksheet['T3'] = "DATE:- 2025-07-09"
        worksheet['Y3'] = "DATE:- 2025-07-10"
        worksheet['C4'] = "TODAY ASSY PLAN"
        worksheet['T4'] = "TOMORROW ASSY PLAN"
        worksheet['Y4'] = "NEXT DAY ASSY PLAN"

    def populate_production_lines(self, worksheet):
        """Populate production line data"""
        
        # CLUTCH ASSY LINE-1 (Rows 7-9)
        clutch_line1_data = [
            {
                'line': 'CLUTCH ASSY LINE-1',
                'a_model': 'KTNA-A500', 'a_plan': 300, 'a_actual': 285, 'a_change': -15, 'a_time': '08:30', 'a_remarks': 'On track',
                'b_model': 'CDN-B200', 'b_plan': 450, 'b_actual': 460, 'b_change': 10, 'b_time': '16:30', 'b_remarks': 'Ahead schedule',
                'c_model': 'KTNA-C100', 'c_plan': 300, 'c_actual': 290, 'c_change': -10, 'c_remarks': 'Minor delay'
            },
            {
                'a_model': 'ABWK-A300', 'a_plan': 500, 'a_actual': 520, 'a_change': 20, 'a_time': '08:45', 'a_remarks': 'Excellent',
                'b_model': 'KILG-B150', 'b_plan': 500, 'b_actual': 485, 'b_change': -15, 'b_time': '16:45', 'b_remarks': 'Material issue',
                'c_model': 'ABWK-C200', 'c_plan': 500, 'c_actual': 495, 'c_change': -5, 'c_remarks': 'Good progress'
            },
            {
                'a_model': 'KTNA-A400', 'a_plan': 300, 'a_actual': 310, 'a_change': 10, 'a_time': '09:00', 'a_remarks': 'Good',
                'b_model': '', 'b_plan': '', 'b_actual': '', 'b_change': '', 'b_time': '', 'b_remarks': '',
                'c_model': 'KTNA-C300', 'c_plan': 300, 'c_actual': 285, 'c_change': -15, 'c_remarks': 'Maintenance delay'
            }
        ]
        
        # Add line name to first row
        worksheet['B7'] = 'CLUTCH ASSY LINE-1'
        
        for i, data in enumerate(clutch_line1_data):
            row = 7 + i
            # A Shift
            worksheet.cell(row=row, column=3, value=data['a_model'])
            worksheet.cell(row=row, column=4, value=data['a_plan'])
            worksheet.cell(row=row, column=5, value=data['a_actual'])
            worksheet.cell(row=row, column=6, value=data['a_change'])
            worksheet.cell(row=row, column=7, value=data['a_time'])
            worksheet.cell(row=row, column=8, value=data['a_remarks'])
            # B Shift
            worksheet.cell(row=row, column=9, value=data['b_model'])
            worksheet.cell(row=row, column=10, value=data['b_plan'])
            worksheet.cell(row=row, column=11, value=data['b_actual'])
            worksheet.cell(row=row, column=12, value=data['b_change'])
            worksheet.cell(row=row, column=13, value=data['b_time'])
            worksheet.cell(row=row, column=14, value=data['b_remarks'])
            # C Shift
            worksheet.cell(row=row, column=15, value=data['c_model'])
            worksheet.cell(row=row, column=16, value=data['c_plan'])
            worksheet.cell(row=row, column=17, value=data['c_actual'])
            worksheet.cell(row=row, column=18, value=data['c_change'])
            worksheet.cell(row=row, column=19, value=data.get('c_time', ''))
            worksheet.cell(row=row, column=20, value=data['c_remarks'])

        # CLUTCH ASSY LINE-2 (Rows 13-15)
        worksheet['B13'] = 'CLUTCH ASSY LINE-2'
        clutch_line2_data = [
            {
                'a_model': 'K1KA-A600', 'a_plan': 500, 'a_actual': 515, 'a_change': 15, 'a_time': '08:15', 'a_remarks': 'Ahead',
                'b_model': 'KILG-B300', 'b_plan': 500, 'b_actual': 490, 'b_change': -10, 'b_time': '16:15', 'b_remarks': 'OK',
                'c_model': 'KTNA-C400', 'c_plan': 300, 'c_actual': 295, 'c_change': -5, 'c_remarks': 'Normal'
            },
            {
                'a_model': 'XFRD-A200', 'a_plan': 400, 'a_actual': 380, 'a_change': -20, 'a_time': '08:30', 'a_remarks': 'Quality check',
                'b_model': 'ABWK-B400', 'b_plan': 500, 'b_actual': 510, 'b_change': 10, 'b_time': '16:30', 'b_remarks': 'Good',
                'c_model': 'ABWK-C300', 'c_plan': 500, 'c_actual': 485, 'c_change': -15, 'c_remarks': 'Tool change'
            },
            {
                'a_model': 'PQRS-A100', 'a_plan': 250, 'a_actual': 260, 'a_change': 10, 'a_time': '08:45', 'a_remarks': 'Efficient',
                'b_model': '', 'b_plan': '', 'b_actual': '', 'b_change': '', 'b_time': '', 'b_remarks': '',
                'c_model': 'KTNA-C500', 'c_plan': 300, 'c_actual': 305, 'c_change': 5, 'c_remarks': 'Excellent'
            }
        ]
        
        for i, data in enumerate(clutch_line2_data):
            row = 13 + i
            # A Shift
            worksheet.cell(row=row, column=3, value=data['a_model'])
            worksheet.cell(row=row, column=4, value=data['a_plan'])
            worksheet.cell(row=row, column=5, value=data['a_actual'])
            worksheet.cell(row=row, column=6, value=data['a_change'])
            worksheet.cell(row=row, column=7, value=data['a_time'])
            worksheet.cell(row=row, column=8, value=data['a_remarks'])
            # B Shift
            worksheet.cell(row=row, column=9, value=data['b_model'])
            worksheet.cell(row=row, column=10, value=data['b_plan'])
            worksheet.cell(row=row, column=11, value=data['b_actual'])
            worksheet.cell(row=row, column=12, value=data['b_change'])
            worksheet.cell(row=row, column=13, value=data['b_time'])
            worksheet.cell(row=row, column=14, value=data['b_remarks'])
            # C Shift
            worksheet.cell(row=row, column=15, value=data['c_model'])
            worksheet.cell(row=row, column=16, value=data['c_plan'])
            worksheet.cell(row=row, column=17, value=data['c_actual'])
            worksheet.cell(row=row, column=18, value=data['c_change'])
            worksheet.cell(row=row, column=20, value=data['c_remarks'])

        # PULLEY ASSY LINE-1 (Rows 19-20)
        worksheet['B19'] = 'PULLEY ASSY LINE-1'
        pulley_data = [
            {
                'a_model': 'XF2Z4-P100', 'a_plan': 2100, 'a_actual': 2080, 'a_change': -20, 'a_time': '08:00', 'a_remarks': 'Minor delay',
                'b_model': 'KILG-P200', 'b_plan': 500, 'b_actual': 520, 'b_change': 20, 'b_time': '16:00', 'b_remarks': 'Good',
                'c_model': 'KTNA-P300', 'c_plan': 300, 'c_actual': 310, 'c_change': 10, 'c_remarks': 'Efficient'
            },
            {
                'a_model': 'PULLEY-A500', 'a_plan': 1500, 'a_actual': 1520, 'a_change': 20, 'a_time': '08:30', 'a_remarks': 'Ahead',
                'b_model': 'ABWK-P100', 'b_plan': 500, 'b_actual': 485, 'b_change': -15, 'b_time': '16:30', 'b_remarks': 'Material delay',
                'c_model': 'PULLEY-C200', 'c_plan': 400, 'c_actual': 395, 'c_change': -5, 'c_remarks': 'Normal'
            }
        ]
        
        for i, data in enumerate(pulley_data):
            row = 19 + i
            # A Shift
            worksheet.cell(row=row, column=3, value=data['a_model'])
            worksheet.cell(row=row, column=4, value=data['a_plan'])
            worksheet.cell(row=row, column=5, value=data['a_actual'])
            worksheet.cell(row=row, column=6, value=data['a_change'])
            worksheet.cell(row=row, column=7, value=data['a_time'])
            worksheet.cell(row=row, column=8, value=data['a_remarks'])
            # B Shift
            worksheet.cell(row=row, column=9, value=data['b_model'])
            worksheet.cell(row=row, column=10, value=data['b_plan'])
            worksheet.cell(row=row, column=11, value=data['b_actual'])
            worksheet.cell(row=row, column=12, value=data['b_change'])
            worksheet.cell(row=row, column=13, value=data['b_time'])
            worksheet.cell(row=row, column=14, value=data['b_remarks'])
            # C Shift
            worksheet.cell(row=row, column=15, value=data['c_model'])
            worksheet.cell(row=row, column=16, value=data['c_plan'])
            worksheet.cell(row=row, column=17, value=data['c_actual'])
            worksheet.cell(row=row, column=18, value=data['c_change'])
            worksheet.cell(row=row, column=20, value=data['c_remarks'])

        # FMD/FFD (Rows 22-23)
        worksheet['B22'] = 'FMD/FFD'
        fmd_data = [
            {
                'a_model': 'FMD-X200', 'a_plan': 800, 'a_actual': 785, 'a_change': -15, 'a_time': '08:20', 'a_remarks': 'Quality focus',
                'b_model': 'FFD-Y300', 'b_plan': 600, 'b_actual': 610, 'b_change': 10, 'b_time': '16:20', 'b_remarks': 'Good',
                'c_model': 'FMD-Z100', 'c_plan': 500, 'c_actual': 495, 'c_change': -5, 'c_remarks': 'OK'
            },
            {
                'a_model': 'FFD-A400', 'a_plan': 1000, 'a_actual': 1015, 'a_change': 15, 'a_time': '08:40', 'a_remarks': 'Excellent',
                'b_model': 'FMD-B500', 'b_plan': 750, 'b_actual': 740, 'b_change': -10, 'b_time': '16:40', 'b_remarks': 'Minor issue',
                'c_model': 'FFD-C600', 'c_plan': 650, 'c_actual': 655, 'c_change': 5, 'c_remarks': 'Good'
            }
        ]
        
        for i, data in enumerate(fmd_data):
            row = 22 + i
            # A Shift
            worksheet.cell(row=row, column=3, value=data['a_model'])
            worksheet.cell(row=row, column=4, value=data['a_plan'])
            worksheet.cell(row=row, column=5, value=data['a_actual'])
            worksheet.cell(row=row, column=6, value=data['a_change'])
            worksheet.cell(row=row, column=7, value=data['a_time'])
            worksheet.cell(row=row, column=8, value=data['a_remarks'])
            # B Shift
            worksheet.cell(row=row, column=9, value=data['b_model'])
            worksheet.cell(row=row, column=10, value=data['b_plan'])
            worksheet.cell(row=row, column=11, value=data['b_actual'])
            worksheet.cell(row=row, column=12, value=data['b_change'])
            worksheet.cell(row=row, column=13, value=data['b_time'])
            worksheet.cell(row=row, column=14, value=data['b_remarks'])
            # C Shift
            worksheet.cell(row=row, column=15, value=data['c_model'])
            worksheet.cell(row=row, column=16, value=data['c_plan'])
            worksheet.cell(row=row, column=17, value=data['c_actual'])
            worksheet.cell(row=row, column=18, value=data['c_change'])
            worksheet.cell(row=row, column=20, value=data['c_remarks'])

        # NEW BUSINESS (Rows 26-27)
        worksheet['B26'] = 'NEW BUSINESS'
        new_business_data = [
            {
                'a_model': 'NB-PROTO1', 'a_plan': 100, 'a_actual': 95, 'a_change': -5, 'a_time': '08:10', 'a_remarks': 'Testing phase',
                'b_model': 'NB-PILOT2', 'b_plan': 150, 'b_actual': 145, 'b_change': -5, 'b_time': '16:10', 'b_remarks': 'Validation',
                'c_model': 'NB-TRIAL3', 'c_plan': 75, 'c_actual': 80, 'c_change': 5, 'c_remarks': 'Good progress'
            },
            {
                'a_model': 'NB-DEV200', 'a_plan': 200, 'a_actual': 205, 'a_change': 5, 'a_time': '08:25', 'a_remarks': 'R&D project',
                'b_model': 'NB-TEST300', 'b_plan': 180, 'b_actual': 175, 'b_change': -5, 'b_time': '16:25', 'b_remarks': 'QA checks',
                'c_model': 'NB-FINAL100', 'c_plan': 120, 'c_actual': 125, 'c_change': 5, 'c_remarks': 'Complete'
            }
        ]
        
        for i, data in enumerate(new_business_data):
            row = 26 + i
            # A Shift
            worksheet.cell(row=row, column=3, value=data['a_model'])
            worksheet.cell(row=row, column=4, value=data['a_plan'])
            worksheet.cell(row=row, column=5, value=data['a_actual'])
            worksheet.cell(row=row, column=6, value=data['a_change'])
            worksheet.cell(row=row, column=7, value=data['a_time'])
            worksheet.cell(row=row, column=8, value=data['a_remarks'])
            # B Shift
            worksheet.cell(row=row, column=9, value=data['b_model'])
            worksheet.cell(row=row, column=10, value=data['b_plan'])
            worksheet.cell(row=row, column=11, value=data['b_actual'])
            worksheet.cell(row=row, column=12, value=data['b_change'])
            worksheet.cell(row=row, column=13, value=data['b_time'])
            worksheet.cell(row=row, column=14, value=data['b_remarks'])
            # C Shift
            worksheet.cell(row=row, column=15, value=data['c_model'])
            worksheet.cell(row=row, column=16, value=data['c_plan'])
            worksheet.cell(row=row, column=17, value=data['c_actual'])
            worksheet.cell(row=row, column=18, value=data['c_change'])
            worksheet.cell(row=row, column=20, value=data['c_remarks'])

    def populate_tomorrow_plans(self, worksheet):
        """Populate tomorrow assembly plans"""
        tomorrow_data = [
            {'model': 'KTNA-T500', 'a_shift': 420, 'b_shift': 380, 'c_shift': 450, 'remarks': 'High priority'},
            {'model': 'COVER-M600', 'a_shift': 600, 'b_shift': 550, 'c_shift': 580, 'remarks': 'Customer demand'},
            {'model': 'ABWK-T300', 'a_shift': 320, 'b_shift': 340, 'c_shift': 310, 'remarks': 'Standard production'},
            {'model': 'XF2Z4-T200', 'a_shift': 250, 'b_shift': 280, 'c_shift': 260, 'remarks': 'Quality focus'},
            {'model': 'PULLEY-T400', 'a_shift': 400, 'b_shift': 420, 'c_shift': 390, 'remarks': 'Increased demand'}
        ]
        
        for i, data in enumerate(tomorrow_data):
            row = 7 + i
            worksheet.cell(row=row, column=21, value=data['model'])    # Column U
            worksheet.cell(row=row, column=22, value=data['a_shift'])  # Column V
            worksheet.cell(row=row, column=23, value=data['b_shift'])  # Column W
            worksheet.cell(row=row, column=24, value=data['c_shift'])  # Column X
            worksheet.cell(row=row, column=25, value=data['remarks'])  # Column Y

    def populate_next_day_plans(self, worksheet):
        """Populate next day assembly plans"""
        next_day_data = [
            {'model': 'KTNA-N600', 'a_shift': 450, 'b_shift': 430, 'c_shift': 470, 'remarks': 'Planning ahead'},
            {'model': 'COVER-N700', 'a_shift': 650, 'b_shift': 600, 'c_shift': 620, 'remarks': 'Bulk order'},
            {'model': 'ABWK-N400', 'a_shift': 380, 'b_shift': 400, 'c_shift': 360, 'remarks': 'Regular production'},
            {'model': 'NEW-MODEL-X', 'a_shift': 200, 'b_shift': 220, 'c_shift': 190, 'remarks': 'New launch'},
            {'model': 'SPECIAL-Y', 'a_shift': 150, 'b_shift': 180, 'c_shift': 160, 'remarks': 'Custom order'}
        ]
        
        for i, data in enumerate(next_day_data):
            row = 7 + i
            worksheet.cell(row=row, column=26, value=data['model'])    # Column Z
            worksheet.cell(row=row, column=27, value=data['a_shift'])  # Column AA
            worksheet.cell(row=row, column=28, value=data['b_shift'])  # Column AB
            worksheet.cell(row=row, column=29, value=data['c_shift'])  # Column AC
            worksheet.cell(row=row, column=30, value=data['remarks'])  # Column AD

    def populate_section_headers(self, worksheet):
        """Populate section headers for parts sections"""
        # Row 29 - Section headers
        worksheet.cell(row=29, column=2, value="CRITICAL PART STATUS")
        worksheet.cell(row=29, column=7, value="AFM PLAN FCIN (MNS)")
        worksheet.cell(row=29, column=11, value="AFM PLAN (I/U)")
        worksheet.cell(row=29, column=27, value="OTHER INFORMATION :")
        
        # Row 30 - Customer headers for SPD
        worksheet.cell(row=30, column=15, value="MSIL")
        worksheet.cell(row=30, column=18, value="HMSI")
        worksheet.cell(row=30, column=21, value="IYM/PIAGGIO")
        worksheet.cell(row=30, column=24, value="HMCL")

    def populate_critical_parts(self, worksheet):
        """Populate critical parts data"""
        critical_parts = [
            {'name': 'Engine Shaft Premium', 'supplier': 'ABC Components Ltd', 'qty': 1200, 'time': '2025-07-10 14:30', 'remarks': 'Critical for Line-1'},
            {'name': 'Gear Box Assembly', 'supplier': 'MDS Technologies', 'qty': 800, 'time': '2025-07-11 09:00', 'remarks': 'Partial delivery expected'},
            {'name': 'Clutch Plate Set', 'supplier': 'Precision Parts Inc', 'qty': 1500, 'time': '2025-07-09 16:00', 'remarks': 'Quality certified'},
            {'name': 'Bearing Housing', 'supplier': 'SKF Bearings', 'qty': 2000, 'time': '2025-07-12 11:30', 'remarks': 'Urgent requirement'},
            {'name': 'Valve Assembly', 'supplier': 'Hydraulic Systems', 'qty': 900, 'time': '2025-07-10 13:00', 'remarks': 'For new project'}
        ]
        
        for i, part in enumerate(critical_parts):
            row = 31 + i
            worksheet.cell(row=row, column=2, value=part['name'])      # Column B
            worksheet.cell(row=row, column=3, value=part['supplier'])  # Column C
            worksheet.cell(row=row, column=4, value=part['qty'])       # Column D
            worksheet.cell(row=row, column=5, value=part['time'])      # Column E
            worksheet.cell(row=row, column=6, value=part['remarks'])   # Column F

    def populate_afm_plans(self, worksheet):
        """Populate AFM plans data"""
        # AFM FCIN plans
        afm_fcin = [
            {'name': 'Motor Controller FCIN', 'number': 'FC-MC-2024-001', 'qty': 500, 'remarks': 'High precision'},
            {'name': 'Sensor Module FCIN', 'number': 'FC-SM-2024-002', 'qty': 750, 'remarks': 'Temperature critical'},
            {'name': 'Drive Unit FCIN', 'number': 'FC-DU-2024-003', 'qty': 300, 'remarks': 'Variable speed'},
            {'name': 'Control Panel FCIN', 'number': 'FC-CP-2024-004', 'qty': 450, 'remarks': 'Touch interface'},
            {'name': 'Safety Switch FCIN', 'number': 'FC-SS-2024-005', 'qty': 600, 'remarks': 'Emergency stop'}
        ]
        
        for i, item in enumerate(afm_fcin):
            row = 31 + i
            worksheet.cell(row=row, column=7, value=item['name'])      # Column G
            worksheet.cell(row=row, column=8, value=item['number'])    # Column H
            worksheet.cell(row=row, column=9, value=item['qty'])       # Column I
            worksheet.cell(row=row, column=10, value=item['remarks'])  # Column J
        
        # AFM I/U plans
        afm_iu = [
            {'name': 'Installation Kit IU', 'number': 'IU-IK-2024-001', 'qty': 200, 'remarks': 'Complete set'},
            {'name': 'Upgrade Module IU', 'number': 'IU-UM-2024-002', 'qty': 350, 'remarks': 'Latest version'},
            {'name': 'Maintenance Tools IU', 'number': 'IU-MT-2024-003', 'qty': 150, 'remarks': 'Specialized tools'},
            {'name': 'Calibration Device IU', 'number': 'IU-CD-2024-004', 'qty': 100, 'remarks': 'Precision required'},
            {'name': 'Testing Equipment IU', 'number': 'IU-TE-2024-005', 'qty': 80, 'remarks': 'Quality assurance'}
        ]
        
        for i, item in enumerate(afm_iu):
            row = 31 + i
            worksheet.cell(row=row, column=11, value=item['name'])     # Column K
            worksheet.cell(row=row, column=12, value=item['number'])   # Column L
            worksheet.cell(row=row, column=13, value=item['qty'])      # Column M
            worksheet.cell(row=row, column=14, value=item['remarks'])  # Column N

    def populate_spd_plans(self, worksheet):
        """Populate SPD plans for all customers"""
        # MSIL
        msil_parts = [
            {'name': 'Swift Engine Mount', 'number': 'MSIL-EM-2024-001', 'qty': 2500, 'remarks': 'High volume'},
            {'name': 'Baleno Brake Disc', 'number': 'MSIL-BD-2024-002', 'qty': 1800, 'remarks': 'Performance grade'},
            {'name': 'Dzire Clutch Set', 'number': 'MSIL-CS-2024-003', 'qty': 2200, 'remarks': 'Premium quality'},
            {'name': 'Ertiga Suspension', 'number': 'MSIL-SU-2024-004', 'qty': 1500, 'remarks': 'Heavy duty'},
            {'name': 'Vitara Steering', 'number': 'MSIL-ST-2024-005', 'qty': 1200, 'remarks': 'Power assisted'}
        ]
        
        for i, part in enumerate(msil_parts):
            row = 31 + i
            worksheet.cell(row=row, column=15, value=part['name'])     # Column O
            worksheet.cell(row=row, column=16, value=part['number'])   # Column P
            worksheet.cell(row=row, column=17, value=part['qty'])      # Column Q
            worksheet.cell(row=row, column=18, value=part['remarks'])  # Column R
        
        # HMSI
        hmsi_parts = [
            {'name': 'Activa CVT Belt', 'number': 'HMSI-CB-2024-001', 'qty': 3000, 'remarks': 'Best seller'},
            {'name': 'City Engine Block', 'number': 'HMSI-EB-2024-002', 'qty': 800, 'remarks': 'Precision casting'},
            {'name': 'Shine Chain Set', 'number': 'HMSI-CH-2024-003', 'qty': 2500, 'remarks': 'Durable chain'},
            {'name': 'Jazz Brake Pad', 'number': 'HMSI-BP-2024-004', 'qty': 1600, 'remarks': 'Ceramic material'},
            {'name': 'Amaze Radiator', 'number': 'HMSI-RA-2024-005', 'qty': 900, 'remarks': 'Cooling system'}
        ]
        
        for i, part in enumerate(hmsi_parts):
            row = 31 + i
            worksheet.cell(row=row, column=19, value=part['name'])     # Column S
            worksheet.cell(row=row, column=20, value=part['number'])   # Column T
            worksheet.cell(row=row, column=21, value=part['qty'])      # Column U
            worksheet.cell(row=row, column=22, value=part['remarks'])  # Column V
        
        # IYM/PIAGGIO
        iym_parts = [
            {'name': 'R15 Engine Piston', 'number': 'IYM-EP-2024-001', 'qty': 1500, 'remarks': 'Racing grade'},
            {'name': 'Vespa Carburetor', 'number': 'PIA-CA-2024-002', 'qty': 800, 'remarks': 'Italian design'},
            {'name': 'FZ Exhaust System', 'number': 'IYM-EX-2024-003', 'qty': 1200, 'remarks': 'Performance tuned'},
            {'name': 'Aprilia Brake Disc', 'number': 'PIA-BD-2024-004', 'qty': 600, 'remarks': 'Sport variant'},
            {'name': 'MT Clutch Assembly', 'number': 'IYM-CL-2024-005', 'qty': 1000, 'remarks': 'Heavy duty'}
        ]
        
        for i, part in enumerate(iym_parts):
            row = 31 + i
            worksheet.cell(row=row, column=23, value=part['name'])     # Column W
            worksheet.cell(row=row, column=24, value=part['number'])   # Column X
            worksheet.cell(row=row, column=25, value=part['qty'])      # Column Y
            worksheet.cell(row=row, column=26, value=part['remarks'])  # Column Z
        
        # HMCL
        hmcl_parts = [
            {'name': 'Splendor Carburetor', 'number': 'HMCL-CA-2024-001', 'qty': 4000, 'remarks': 'Economy segment'},
            {'name': 'Passion Engine Kit', 'number': 'HMCL-EK-2024-002', 'qty': 3500, 'remarks': 'Complete overhaul'},
            {'name': 'Glamour Brake Set', 'number': 'HMCL-BS-2024-003', 'qty': 2800, 'remarks': 'Safety critical'},
            {'name': 'HF Dawn Chain', 'number': 'HMCL-CH-2024-004', 'qty': 3200, 'remarks': 'Commercial use'},
            {'name': 'Super Splendor Piston', 'number': 'HMCL-PI-2024-005', 'qty': 2600, 'remarks': 'Premium grade'}
        ]
        
        for i, part in enumerate(hmcl_parts):
            row = 31 + i
            worksheet.cell(row=row, column=27, value=part['name'])     # Column AA
            worksheet.cell(row=row, column=28, value=part['number'])   # Column AB
            worksheet.cell(row=row, column=29, value=part['qty'])      # Column AC
            worksheet.cell(row=row, column=30, value=part['remarks'])  # Column AD

    def populate_other_information(self, worksheet):
        """Populate other information data"""
        other_info = [
            {'name': 'Special Tool Set', 'qty': 50, 'date': '2025-07-15', 'remarks': 'Maintenance dept'},
            {'name': 'Testing Jigs', 'qty': 25, 'date': '2025-07-20', 'remarks': 'Quality control'},
            {'name': 'Calibration Weights', 'qty': 100, 'date': '2025-07-12', 'remarks': 'Precision required'},
            {'name': 'Safety Equipment', 'qty': 200, 'date': '2025-07-18', 'remarks': 'Worker protection'},
            {'name': 'Cleaning Supplies', 'qty': 500, 'date': '2025-07-25', 'remarks': 'Monthly stock'}
        ]
        
        for i, item in enumerate(other_info):
            row = 31 + i
            worksheet.cell(row=row, column=31, value=item['name'])     # Column AE
            worksheet.cell(row=row, column=32, value=item['qty'])      # Column AF
            worksheet.cell(row=row, column=33, value=item['date'])     # Column AG
            worksheet.cell(row=row, column=34, value=item['remarks'])  # Column AH